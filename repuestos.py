from pathlib import Path
import sqlite3
import unicodedata
import shutil
import zipfile
from copy import deepcopy
from datetime import datetime, date
from xml.etree import ElementTree as ET

from flask import (
    Blueprint,
    render_template,
    request,
    jsonify,
    redirect,
    url_for,
    flash,
    send_file,
)


repuestos_bp = Blueprint(
    "repuestos",
    __name__,
    url_prefix="/repuestos",
)


BASE_DIR = Path(__file__).resolve().parent
CARPETA_DATOS = BASE_DIR / "datos"
RUTA_BD = CARPETA_DATOS / "repuestos.db"


def conectar_bd():
    CARPETA_DATOS.mkdir(
        parents=True,
        exist_ok=True,
    )

    conexion = sqlite3.connect(
        RUTA_BD,
        timeout=60,
    )

    conexion.row_factory = sqlite3.Row

    conexion.execute(
        "PRAGMA foreign_keys = ON"
    )

    conexion.execute(
        "PRAGMA busy_timeout = 60000"
    )

    return conexion


def cargar_openpyxl():
    try:
        from openpyxl import load_workbook
        return load_workbook

    except ModuleNotFoundError as error:
        raise RuntimeError(
            "Falta instalar openpyxl. "
            "Ejecute: python -m pip install openpyxl"
        ) from error


def normalizar_encabezado(valor) -> str:
    texto = str(
        valor or ""
    ).strip().upper()

    texto = "".join(
        caracter
        for caracter in unicodedata.normalize(
            "NFD",
            texto,
        )
        if unicodedata.category(
            caracter
        ) != "Mn"
    )

    return " ".join(
        texto.split()
    )


def convertir_numero(valor) -> float:
    if valor in (None, ""):
        return 0.0

    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()

    if not texto:
        return 0.0

    texto = (
        texto
        .replace("$", "")
        .replace("COP", "")
        .replace(" ", "")
    )

    try:
        return float(texto)
    except ValueError:
        pass

    try:
        if "." in texto and "," in texto:
            texto = (
                texto
                .replace(".", "")
                .replace(",", ".")
            )
        elif "," in texto:
            texto = texto.replace(",", ".")

        return float(texto)

    except ValueError:
        return 0.0


def normalizar_codigo_sap(valor) -> str:
    if valor in (None, ""):
        return ""

    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))

    return str(valor).strip()


def inicializar_bd():
    conexion = conectar_bd()

    try:
        conexion.executescript(
            """
            CREATE TABLE IF NOT EXISTS materiales_sap (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo_sap TEXT UNIQUE,
                descripcion TEXT NOT NULL,
                unidad_medida TEXT,
                existencia REAL DEFAULT 0,
                valor_inventario REAL DEFAULT 0,
                valor_unitario REAL DEFAULT 0,
                origen_dato TEXT DEFAULT 'SAP',
                fecha_actualizacion TEXT,
                activo INTEGER DEFAULT 1
            );

            CREATE TABLE IF NOT EXISTS solicitudes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo_solicitud TEXT UNIQUE NOT NULL,
                semana INTEGER NOT NULL,
                anio INTEGER NOT NULL,
                fecha_creacion TEXT NOT NULL,
                area TEXT DEFAULT 'MP TAPA PLASTICA',
                estado TEXT DEFAULT 'GENERADA',
                archivo_generado TEXT
            );

            CREATE TABLE IF NOT EXISTS solicitud_detalle (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                solicitud_id INTEGER NOT NULL,
                codigo_sap TEXT,
                descripcion TEXT NOT NULL,
                origen TEXT,
                unidad_medida TEXT,
                existencia_al_solicitar REAL DEFAULT 0,
                consumo_promedio REAL DEFAULT 0,
                tiempo_entrega REAL DEFAULT 0,
                cantidad_pedir REAL DEFAULT 0,
                valor_unitario REAL DEFAULT 0,
                observaciones TEXT,
                origen_dato TEXT DEFAULT 'SAP',
                estado TEXT DEFAULT 'PENDIENTE',
                cantidad_recibida REAL DEFAULT 0,
                cantidad_retirada REAL DEFAULT 0,
                fecha_recibido TEXT,
                fecha_retiro TEXT,
                FOREIGN KEY (
                    solicitud_id
                )
                REFERENCES solicitudes(id)
                ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS secciones (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT UNIQUE NOT NULL,
                descripcion TEXT,
                activo INTEGER DEFAULT 1,
                fecha_creacion TEXT
            );

            CREATE TABLE IF NOT EXISTS subsecciones (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                seccion_id INTEGER NOT NULL,
                nombre TEXT NOT NULL,
                descripcion TEXT,
                activo INTEGER DEFAULT 1,
                fecha_creacion TEXT,
                UNIQUE (seccion_id, nombre),
                FOREIGN KEY (seccion_id)
                REFERENCES secciones(id)
                ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS inventario_tecnico (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo_sap TEXT,
                descripcion TEXT NOT NULL,
                unidad_medida TEXT,
                seccion_id INTEGER NOT NULL,
                subseccion_id INTEGER,
                tipo_control TEXT DEFAULT 'manual',
                stock_minimo REAL DEFAULT 0,
                stock_objetivo REAL DEFAULT 0,
                factor_minimo REAL DEFAULT 1,
                factor_objetivo REAL DEFAULT 1,
                origen_dato TEXT DEFAULT 'SAP',
                observaciones TEXT,
                activo INTEGER DEFAULT 1,
                fecha_creacion TEXT,
                FOREIGN KEY (seccion_id)
                REFERENCES secciones(id),
                FOREIGN KEY (subseccion_id)
                REFERENCES subsecciones(id)
            );

            CREATE TABLE IF NOT EXISTS inventario_maquinas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                inventario_id INTEGER NOT NULL,
                maquina_id INTEGER NOT NULL,
                cantidad_necesaria REAL DEFAULT 0,
                FOREIGN KEY (inventario_id)
                REFERENCES inventario_tecnico(id)
                ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS recepciones (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                detalle_solicitud_id INTEGER NOT NULL,
                cantidad REAL NOT NULL,
                fecha_recepcion TEXT NOT NULL,
                confirmado INTEGER DEFAULT 0,
                observaciones TEXT,
                FOREIGN KEY (detalle_solicitud_id)
                REFERENCES solicitud_detalle(id)
                ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS retiros (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                detalle_solicitud_id INTEGER NOT NULL,
                cantidad REAL NOT NULL,
                fecha_retiro TEXT NOT NULL,
                observaciones TEXT,
                FOREIGN KEY (detalle_solicitud_id)
                REFERENCES solicitud_detalle(id)
                ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS snapshots_sap (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha_carga TEXT NOT NULL,
                codigo_sap TEXT NOT NULL,
                descripcion TEXT,
                unidad_medida TEXT,
                existencia REAL DEFAULT 0,
                valor_inventario REAL DEFAULT 0,
                valor_unitario REAL DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS cargas_sap (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha_carga TEXT NOT NULL,
                nombre_archivo TEXT,
                cantidad_registros INTEGER DEFAULT 0
            );
            """
        )

        conexion.commit()

    finally:
        conexion.close()


def obtener_resumen():
    conexion = conectar_bd()

    try:
        solicitudes_activas = conexion.execute(
            """
            SELECT COUNT(*) AS total
            FROM solicitudes
            WHERE estado NOT IN ('CERRADA', 'ANULADA')
            """
        ).fetchone()["total"]

        pendientes_recibir = conexion.execute(
            """
            SELECT COUNT(*) AS total
            FROM solicitud_detalle
            WHERE estado IN ('PENDIENTE', 'RECIBIDO_PARCIAL')
            """
        ).fetchone()["total"]

        recibidos_retirar = conexion.execute(
            """
            SELECT COUNT(*) AS total
            FROM solicitud_detalle
            WHERE estado IN ('RECIBIDO', 'PENDIENTE_RETIRO')
            AND cantidad_recibida > cantidad_retirada
            """
        ).fetchone()["total"]

        bajo_stock = conexion.execute(
            """
            SELECT COUNT(*) AS total
            FROM inventario_tecnico inventario
            LEFT JOIN materiales_sap material
                ON material.codigo_sap = inventario.codigo_sap
            WHERE inventario.activo = 1
            AND COALESCE(material.existencia, 0) < inventario.stock_minimo
            """
        ).fetchone()["total"]

        return {
            "solicitudes_activas": int(solicitudes_activas or 0),
            "pendientes_recibir": int(pendientes_recibir or 0),
            "recibidos_retirar": int(recibidos_retirar or 0),
            "bajo_stock": int(bajo_stock or 0),
        }

    finally:
        conexion.close()


def buscar_materiales(texto: str):
    texto = texto.strip().upper()

    if not texto:
        return []

    conexion = conectar_bd()

    try:
        filas = conexion.execute(
            """
            SELECT
                id,
                codigo_sap,
                descripcion,
                unidad_medida,
                existencia,
                valor_inventario,
                valor_unitario,
                origen_dato
            FROM materiales_sap
            WHERE activo = 1
            AND (
                UPPER(COALESCE(codigo_sap, '')) LIKE ?
                OR UPPER(descripcion) LIKE ?
            )
            ORDER BY
                CASE
                    WHEN UPPER(COALESCE(codigo_sap, '')) = ?
                    THEN 0
                    ELSE 1
                END,
                descripcion
            LIMIT 40
            """,
            (
                f"%{texto}%",
                f"%{texto}%",
                texto,
            ),
        ).fetchall()

        return [dict(fila) for fila in filas]

    finally:
        conexion.close()


def buscar_pedidos_pendientes(codigo_sap: str):
    if not codigo_sap:
        return []

    conexion = conectar_bd()

    try:
        filas = conexion.execute(
            """
            SELECT
                sd.id,
                sd.codigo_sap,
                sd.descripcion,
                sd.cantidad_pedir,
                sd.cantidad_recibida,
                sd.estado,
                s.codigo_solicitud,
                s.semana,
                s.anio,
                s.fecha_creacion
            FROM solicitud_detalle sd
            INNER JOIN solicitudes s
                ON s.id = sd.solicitud_id
            WHERE sd.codigo_sap = ?
            AND sd.estado IN (
                'PENDIENTE',
                'RECIBIDO_PARCIAL'
            )
            AND s.estado NOT IN ('CERRADA', 'ANULADA')
            ORDER BY sd.id DESC
            """,
            (codigo_sap,),
        ).fetchall()

        resultado = []

        for fila in filas:
            item = dict(fila)
            item["cantidad_pendiente"] = max(
                0,
                float(item["cantidad_pedir"] or 0)
                - float(item["cantidad_recibida"] or 0),
            )
            resultado.append(item)

        return resultado

    finally:
        conexion.close()


def siguiente_codigo_solicitud(
    conexion,
    anio: int,
    semana: int,
) -> str:
    prefijo = f"SR-{anio}-{semana:02d}-"

    fila = conexion.execute(
        """
        SELECT codigo_solicitud
        FROM solicitudes
        WHERE codigo_solicitud LIKE ?
        ORDER BY id DESC
        LIMIT 1
        """,
        (f"{prefijo}%",),
    ).fetchone()

    consecutivo = 1

    if fila:
        try:
            consecutivo = int(
                str(
                    fila["codigo_solicitud"]
                ).split("-")[-1]
            ) + 1

        except (
            TypeError,
            ValueError,
            IndexError,
        ):
            consecutivo = 1

    return (
        f"{prefijo}"
        f"{consecutivo:03d}"
    )


def validar_item_solicitud(item: dict):
    descripcion = str(
        item.get(
            "descripcion",
            ""
        )
    ).strip()

    unidad = str(
        item.get(
            "unidad_medida",
            ""
        )
    ).strip()

    origen = str(
        item.get(
            "origen",
            ""
        )
    ).strip().upper()

    if not descripcion:
        raise ValueError(
            "Todos los ítems deben tener descripción."
        )

    if not unidad:
        raise ValueError(
            f"El material '{descripcion}' no tiene unidad de medida."
        )

    if origen not in (
        "NAL",
        "IMP",
    ):
        raise ValueError(
            f"Seleccione el origen NAL o IMP para '{descripcion}'."
        )

    cantidad = convertir_numero(
        item.get(
            "cantidad_pedir"
        )
    )

    if cantidad <= 0:
        raise ValueError(
            f"La cantidad a pedir de '{descripcion}' debe ser mayor que cero."
        )

    consumo = convertir_numero(
        item.get(
            "consumo_promedio"
        )
    )

    tiempo_entrega = convertir_numero(
        item.get(
            "tiempo_entrega"
        )
    )

    valor_unitario = convertir_numero(
        item.get(
            "valor_unitario"
        )
    )

    return {
        "codigo_sap":
            normalizar_codigo_sap(
                item.get(
                    "codigo_sap"
                )
            ),

        "descripcion":
            descripcion,

        "origen":
            origen,

        "unidad_medida":
            unidad,

        "existencia_al_solicitar":
            convertir_numero(
                item.get(
                    "existencia"
                )
            ),

        "consumo_promedio":
            consumo,

        "tiempo_entrega":
            tiempo_entrega,

        "cantidad_pedir":
            cantidad,

        "valor_unitario":
            valor_unitario,

        "observaciones":
            str(
                item.get(
                    "observaciones",
                    ""
                )
            ).strip(),

        "origen_dato":
            str(
                item.get(
                    "origen_dato",
                    "SAP"
                )
            ).strip().upper()
            or "SAP",
    }


def crear_solicitud(
    items: list[dict],
) -> dict:
    if not items:
        raise ValueError(
            "La solicitud no contiene materiales."
        )

    items_validados = [
        validar_item_solicitud(
            item
        )
        for item in items
    ]

    codigos = [
        item[
            "codigo_sap"
        ]
        for item in items_validados
        if item[
            "codigo_sap"
        ]
    ]

    if len(
        codigos
    ) != len(
        set(
            codigos
        )
    ):
        raise ValueError(
            "El pedido contiene códigos SAP repetidos."
        )

    ahora = datetime.now()
    calendario = ahora.isocalendar()

    anio = int(
        calendario.year
    )

    semana = int(
        calendario.week
    )

    fecha_creacion = ahora.isoformat(
        timespec="seconds"
    )

    conexion = conectar_bd()

    try:
        codigo_solicitud = siguiente_codigo_solicitud(
            conexion,
            anio,
            semana,
        )

        cursor = conexion.execute(
            """
            INSERT INTO solicitudes (
                codigo_solicitud,
                semana,
                anio,
                fecha_creacion,
                area,
                estado
            )
            VALUES (
                ?,
                ?,
                ?,
                ?,
                'MP TAPA PLASTICA',
                'GENERADA'
            )
            """,
            (
                codigo_solicitud,
                semana,
                anio,
                fecha_creacion,
            ),
        )

        solicitud_id = cursor.lastrowid

        filas_detalle = []

        for item in items_validados:
            filas_detalle.append(
                (
                    solicitud_id,
                    item[
                        "codigo_sap"
                    ],
                    item[
                        "descripcion"
                    ],
                    item[
                        "origen"
                    ],
                    item[
                        "unidad_medida"
                    ],
                    item[
                        "existencia_al_solicitar"
                    ],
                    item[
                        "consumo_promedio"
                    ],
                    item[
                        "tiempo_entrega"
                    ],
                    item[
                        "cantidad_pedir"
                    ],
                    item[
                        "valor_unitario"
                    ],
                    item[
                        "observaciones"
                    ],
                    item[
                        "origen_dato"
                    ],
                    "PENDIENTE",
                )
            )

        conexion.executemany(
            """
            INSERT INTO solicitud_detalle (
                solicitud_id,
                codigo_sap,
                descripcion,
                origen,
                unidad_medida,
                existencia_al_solicitar,
                consumo_promedio,
                tiempo_entrega,
                cantidad_pedir,
                valor_unitario,
                observaciones,
                origen_dato,
                estado
            )
            VALUES (
                ?,
                ?,
                ?,
                ?,
                ?,
                ?,
                ?,
                ?,
                ?,
                ?,
                ?,
                ?,
                ?
            )
            """,
            filas_detalle,
        )

        conexion.commit()

        return {
            "id":
                solicitud_id,

            "codigo_solicitud":
                codigo_solicitud,

            "semana":
                semana,

            "anio":
                anio,

            "fecha_creacion":
                fecha_creacion,

            "cantidad_items":
                len(
                    items_validados
                ),

            "valor_estimado":
                sum(
                    item[
                        "cantidad_pedir"
                    ]
                    * item[
                        "valor_unitario"
                    ]
                    for item in items_validados
                ),
        }

    except Exception:
        conexion.rollback()
        raise

    finally:
        conexion.close()



NS_XLSX = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
ET.register_namespace("", NS_XLSX)


def qn(nombre: str) -> str:
    return f"{{{NS_XLSX}}}{nombre}"


def carpeta_formatos() -> Path:
    return BASE_DIR / "formatos"


def carpeta_salidas_repuestos() -> Path:
    ruta = BASE_DIR / "salidas" / "repuestos"
    ruta.mkdir(
        parents=True,
        exist_ok=True,
    )
    return ruta


def localizar_plantilla_repuestos() -> Path:
    formatos = carpeta_formatos()

    if not formatos.exists():
        raise FileNotFoundError(
            "No existe la carpeta 'formatos'."
        )

    candidatos = []

    for ruta in formatos.glob("*.xlsx"):
        nombre = normalizar_encabezado(
            ruta.stem
        )

        if (
            "SOLICITUD" in nombre
            and "REPUEST" in nombre
        ):
            candidatos.append(
                ruta
            )

    if not candidatos:
        raise FileNotFoundError(
            "No se encontró la plantilla corporativa de Solicitud de Repuestos "
            "en la carpeta 'formatos'."
        )

    candidatos.sort(
        key=lambda ruta: ruta.stat().st_mtime,
        reverse=True,
    )

    return candidatos[0]


def nombre_salida_semanal(
    semana: int,
) -> Path:
    carpeta = carpeta_salidas_repuestos()

    base = (
        f"SOLICITUD REPUESTOS SEMANA {semana}.xlsx"
    )

    ruta = carpeta / base

    if not ruta.exists():
        return ruta

    consecutivo = 2

    while True:
        ruta = carpeta / (
            f"SOLICITUD REPUESTOS SEMANA "
            f"{semana}_{consecutivo:02d}.xlsx"
        )

        if not ruta.exists():
            return ruta

        consecutivo += 1


def fecha_a_serial_excel(
    valor: date,
) -> int:
    origen = date(
        1899,
        12,
        30,
    )

    return (
        valor
        - origen
    ).days


def obtener_o_crear_celda(
    fila_xml,
    referencia: str,
    estilo: str | None = None,
):
    for celda in fila_xml.findall(
        qn("c")
    ):
        if celda.get(
            "r"
        ) == referencia:
            return celda

    celda = ET.Element(
        qn("c"),
        {
            "r":
                referencia
        },
    )

    if estilo is not None:
        celda.set(
            "s",
            estilo,
        )

    fila_xml.append(
        celda
    )

    return celda


def limpiar_contenido_celda(
    celda,
):
    for hijo in list(
        celda
    ):
        celda.remove(
            hijo
        )

    celda.attrib.pop(
        "t",
        None,
    )


def escribir_texto_xml(
    celda,
    valor,
):
    limpiar_contenido_celda(
        celda
    )

    celda.set(
        "t",
        "inlineStr",
    )

    is_node = ET.SubElement(
        celda,
        qn("is"),
    )

    t_node = ET.SubElement(
        is_node,
        qn("t"),
    )

    texto = str(
        valor
        if valor is not None
        else ""
    )

    if (
        texto.startswith(" ")
        or texto.endswith(" ")
        or "\n" in texto
    ):
        t_node.set(
            "{http://www.w3.org/XML/1998/namespace}space",
            "preserve",
        )

    t_node.text = texto


def escribir_numero_xml(
    celda,
    valor,
):
    limpiar_contenido_celda(
        celda
    )

    numero = float(
        valor or 0
    )

    if numero.is_integer():
        texto = str(
            int(
                numero
            )
        )
    else:
        texto = (
            f"{numero:.10f}"
            .rstrip("0")
            .rstrip(".")
        )

    v_node = ET.SubElement(
        celda,
        qn("v"),
    )

    v_node.text = texto


def escribir_formula_xml(
    celda,
    formula: str,
):
    limpiar_contenido_celda(
        celda
    )

    f_node = ET.SubElement(
        celda,
        qn("f"),
    )

    f_node.text = formula

    v_node = ET.SubElement(
        celda,
        qn("v"),
    )

    v_node.text = "0"


def columna_de_referencia(
    referencia: str,
) -> str:
    resultado = []

    for caracter in referencia:
        if caracter.isalpha():
            resultado.append(
                caracter
            )
        else:
            break

    return "".join(
        resultado
    )


def duplicar_fila_plantilla(
    sheet_data,
    fila_base,
    numero_fila: int,
):
    nueva = deepcopy(
        fila_base
    )

    nueva.set(
        "r",
        str(
            numero_fila
        ),
    )

    nueva.attrib.pop(
        "spans",
        None,
    )

    for celda in nueva.findall(
        qn("c")
    ):
        referencia_original = celda.get(
            "r",
            ""
        )

        columna = columna_de_referencia(
            referencia_original
        )

        celda.set(
            "r",
            f"{columna}{numero_fila}",
        )

        limpiar_contenido_celda(
            celda
        )

    sheet_data.append(
        nueva
    )

    return nueva


def ordenar_filas_sheet_data(
    sheet_data,
):
    filas = list(
        sheet_data.findall(
            qn("row")
        )
    )

    filas.sort(
        key=lambda fila: int(
            fila.get(
                "r",
                "0",
            )
        )
    )

    for fila in filas:
        sheet_data.remove(
            fila
        )

    for fila in filas:
        sheet_data.append(
            fila
        )


def actualizar_dimension_hoja(
    root,
    ultima_fila: int,
):
    dimension = root.find(
        qn("dimension")
    )

    if dimension is not None:
        dimension.set(
            "ref",
            f"A1:Z{ultima_fila}",
        )


def actualizar_tabla_excel(
    archivos_zip: dict[str, bytes],
    ultima_fila: int,
):
    ruta_tabla = "xl/tables/table1.xml"

    if ruta_tabla not in archivos_zip:
        return

    root = ET.fromstring(
        archivos_zip[
            ruta_tabla
        ]
    )

    root.set(
        "ref",
        f"A10:N{ultima_fila}",
    )

    auto_filter = root.find(
        qn("autoFilter")
    )

    if auto_filter is not None:
        auto_filter.set(
            "ref",
            f"A10:N{ultima_fila}",
        )

    archivos_zip[
        ruta_tabla
    ] = ET.tostring(
        root,
        encoding="utf-8",
        xml_declaration=True,
    )


def obtener_solicitud_completa(
    solicitud_id: int,
):
    conexion = conectar_bd()

    try:
        cabecera = conexion.execute(
            """
            SELECT
                id,
                codigo_solicitud,
                semana,
                anio,
                fecha_creacion,
                area,
                estado,
                archivo_generado
            FROM solicitudes
            WHERE id = ?
            """,
            (
                solicitud_id,
            ),
        ).fetchone()

        if cabecera is None:
            raise ValueError(
                "La solicitud no existe."
            )

        detalles = conexion.execute(
            """
            SELECT
                id,
                codigo_sap,
                descripcion,
                origen,
                unidad_medida,
                existencia_al_solicitar,
                consumo_promedio,
                tiempo_entrega,
                cantidad_pedir,
                valor_unitario,
                observaciones,
                origen_dato
            FROM solicitud_detalle
            WHERE solicitud_id = ?
            ORDER BY id
            """,
            (
                solicitud_id,
            ),
        ).fetchall()

        return (
            dict(
                cabecera
            ),
            [
                dict(
                    fila
                )
                for fila in detalles
            ],
        )

    finally:
        conexion.close()


def generar_excel_corporativo(
    solicitud_id: int,
) -> Path:
    solicitud, detalles = (
        obtener_solicitud_completa(
            solicitud_id
        )
    )

    if not detalles:
        raise ValueError(
            "La solicitud no tiene ítems para generar."
        )

    plantilla = localizar_plantilla_repuestos()

    salida = nombre_salida_semanal(
        int(
            solicitud[
                "semana"
            ]
        )
    )

    with zipfile.ZipFile(
        plantilla,
        "r",
    ) as origen_zip:
        archivos = {
            nombre:
                origen_zip.read(
                    nombre
                )
            for nombre in origen_zip.namelist()
        }

    ruta_sheet = (
        "xl/worksheets/sheet1.xml"
    )

    if ruta_sheet not in archivos:
        raise ValueError(
            "La plantilla no contiene la hoja corporativa esperada."
        )

    root = ET.fromstring(
        archivos[
            ruta_sheet
        ]
    )

    sheet_data = root.find(
        qn("sheetData")
    )

    if sheet_data is None:
        raise ValueError(
            "No se encontró sheetData en la plantilla."
        )

    filas_por_numero = {
        int(
            fila.get(
                "r",
                "0",
            )
        ):
            fila
        for fila in sheet_data.findall(
            qn("row")
        )
    }

    fila_base = filas_por_numero.get(
        21
    )

    if fila_base is None:
        fila_base = filas_por_numero.get(
            11
        )

    if fila_base is None:
        raise ValueError(
            "No se encontró una fila de detalle utilizable en la plantilla."
        )

    primera_fila = 11

    ultima_fila_necesaria = max(
        21,
        primera_fila
        + len(
            detalles
        )
        - 1,
    )

    for numero_fila in range(
        11,
        ultima_fila_necesaria + 1,
    ):
        if numero_fila not in filas_por_numero:
            nueva = duplicar_fila_plantilla(
                sheet_data,
                fila_base,
                numero_fila,
            )

            filas_por_numero[
                numero_fila
            ] = nueva

    ordenar_filas_sheet_data(
        sheet_data
    )

    # Fecha corporativa en B6:J6.
    fila6 = filas_por_numero.get(
        6
    )

    if fila6 is not None:
        celda_b6 = obtener_o_crear_celda(
            fila6,
            "B6",
        )

        escribir_numero_xml(
            celda_b6,
            fecha_a_serial_excel(
                datetime.now().date()
            ),
        )

    # Área corporativa en N6.
    if fila6 is not None:
        celda_n6 = obtener_o_crear_celda(
            fila6,
            "N6",
        )

        escribir_texto_xml(
            celda_n6,
            "MP TAPA PLASTICA",
        )

    for indice, numero_fila in enumerate(
        range(
            11,
            ultima_fila_necesaria + 1,
        )
    ):
        fila_xml = filas_por_numero[
            numero_fila
        ]

        if indice < len(
            detalles
        ):
            item = detalles[
                indice
            ]

            codigo = (
                item.get(
                    "codigo_sap"
                )
                or "CREAR"
            )

            escribir_texto_xml(
                obtener_o_crear_celda(
                    fila_xml,
                    f"A{numero_fila}",
                ),
                "MP TAPA PLASTICA",
            )

            # COD SAP se escribe como texto para preservar códigos.
            escribir_texto_xml(
                obtener_o_crear_celda(
                    fila_xml,
                    f"B{numero_fila}",
                ),
                codigo,
            )

            escribir_texto_xml(
                obtener_o_crear_celda(
                    fila_xml,
                    f"C{numero_fila}",
                ),
                item.get(
                    "descripcion",
                    "",
                ),
            )

            escribir_texto_xml(
                obtener_o_crear_celda(
                    fila_xml,
                    f"D{numero_fila}",
                ),
                item.get(
                    "origen",
                    "",
                ),
            )

            escribir_texto_xml(
                obtener_o_crear_celda(
                    fila_xml,
                    f"E{numero_fila}",
                ),
                item.get(
                    "unidad_medida",
                    "",
                ),
            )

            escribir_numero_xml(
                obtener_o_crear_celda(
                    fila_xml,
                    f"F{numero_fila}",
                ),
                item.get(
                    "existencia_al_solicitar",
                    0,
                ),
            )

            escribir_numero_xml(
                obtener_o_crear_celda(
                    fila_xml,
                    f"G{numero_fila}",
                ),
                item.get(
                    "consumo_promedio",
                    0,
                ),
            )

            escribir_formula_xml(
                obtener_o_crear_celda(
                    fila_xml,
                    f"H{numero_fila}",
                ),
                (
                    f"IFERROR("
                    f"F{numero_fila}/"
                    f"G{numero_fila},0)"
                ),
            )

            escribir_numero_xml(
                obtener_o_crear_celda(
                    fila_xml,
                    f"I{numero_fila}",
                ),
                item.get(
                    "tiempo_entrega",
                    0,
                ),
            )

            escribir_numero_xml(
                obtener_o_crear_celda(
                    fila_xml,
                    f"J{numero_fila}",
                ),
                item.get(
                    "cantidad_pedir",
                    0,
                ),
            )

            escribir_formula_xml(
                obtener_o_crear_celda(
                    fila_xml,
                    f"K{numero_fila}",
                ),
                (
                    f"IFERROR(("
                    f"F{numero_fila}+"
                    f"J{numero_fila})/"
                    f"G{numero_fila},0)"
                ),
            )

            escribir_numero_xml(
                obtener_o_crear_celda(
                    fila_xml,
                    f"L{numero_fila}",
                ),
                item.get(
                    "valor_unitario",
                    0,
                ),
            )

            escribir_formula_xml(
                obtener_o_crear_celda(
                    fila_xml,
                    f"M{numero_fila}",
                ),
                (
                    f"L{numero_fila}*"
                    f"J{numero_fila}"
                ),
            )

            escribir_texto_xml(
                obtener_o_crear_celda(
                    fila_xml,
                    f"N{numero_fila}",
                ),
                item.get(
                    "observaciones",
                    "",
                ),
            )

        else:
            # Limpia filas sobrantes conservando estilos.
            for columna in (
                "A",
                "B",
                "C",
                "D",
                "E",
                "F",
                "G",
                "I",
                "J",
                "L",
                "N",
            ):
                celda = obtener_o_crear_celda(
                    fila_xml,
                    f"{columna}{numero_fila}",
                )

                limpiar_contenido_celda(
                    celda
                )

            escribir_formula_xml(
                obtener_o_crear_celda(
                    fila_xml,
                    f"H{numero_fila}",
                ),
                (
                    f"IFERROR("
                    f"F{numero_fila}/"
                    f"G{numero_fila},0)"
                ),
            )

            escribir_formula_xml(
                obtener_o_crear_celda(
                    fila_xml,
                    f"K{numero_fila}",
                ),
                (
                    f"IFERROR(("
                    f"F{numero_fila}+"
                    f"J{numero_fila})/"
                    f"G{numero_fila},0)"
                ),
            )

            escribir_formula_xml(
                obtener_o_crear_celda(
                    fila_xml,
                    f"M{numero_fila}",
                ),
                (
                    f"L{numero_fila}*"
                    f"J{numero_fila}"
                ),
            )

    actualizar_dimension_hoja(
        root,
        ultima_fila_necesaria,
    )

    archivos[
        ruta_sheet
    ] = ET.tostring(
        root,
        encoding="utf-8",
        xml_declaration=True,
    )

    actualizar_tabla_excel(
        archivos,
        ultima_fila_necesaria,
    )

    with zipfile.ZipFile(
        salida,
        "w",
        zipfile.ZIP_DEFLATED,
    ) as destino_zip:
        for nombre, contenido in archivos.items():
            destino_zip.writestr(
                nombre,
                contenido,
            )

    conexion = conectar_bd()

    try:
        conexion.execute(
            """
            UPDATE solicitudes
            SET archivo_generado = ?
            WHERE id = ?
            """,
            (
                str(
                    salida
                ),
                solicitud_id,
            ),
        )

        conexion.commit()

    finally:
        conexion.close()

    return salida


def encontrar_fila_encabezados(hoja) -> int:
    requeridos = {
        "MATERIAL",
        "TEXTO BREVE DE MATERIAL",
    }

    limite = min(
        hoja.max_row,
        30,
    )

    for fila in range(
        1,
        limite + 1,
    ):
        encontrados = set()

        for columna in range(
            1,
            hoja.max_column + 1,
        ):
            encabezado = normalizar_encabezado(
                hoja.cell(
                    fila,
                    columna,
                ).value
            )

            if encabezado:
                encontrados.add(
                    encabezado
                )

        if requeridos.issubset(
            encontrados
        ):
            return fila

    raise ValueError(
        "No se encontró una fila de encabezados válida. "
        "El archivo debe contener como mínimo 'Material' "
        "y 'Texto breve de material'."
    )


def obtener_mapa_columnas(
    hoja,
    fila_encabezado: int,
) -> dict[str, int]:
    mapa = {}

    for columna in range(
        1,
        hoja.max_column + 1,
    ):
        encabezado = normalizar_encabezado(
            hoja.cell(
                fila_encabezado,
                columna,
            ).value
        )

        if encabezado:
            mapa[
                encabezado
            ] = columna

    return mapa


def buscar_columna(
    mapa: dict[str, int],
    opciones: list[str],
):
    for opcion in opciones:
        normalizada = normalizar_encabezado(
            opcion
        )

        if normalizada in mapa:
            return mapa[
                normalizada
            ]

    return None


def importar_base_sap(
    ruta_excel: Path,
) -> dict:
    load_workbook = cargar_openpyxl()

    workbook = load_workbook(
        ruta_excel,
        data_only=True,
        read_only=True,
    )

    try:
        hoja = workbook[
            workbook.sheetnames[0]
        ]

        fila_encabezado = encontrar_fila_encabezados(
            hoja
        )

        encabezados = obtener_mapa_columnas(
            hoja,
            fila_encabezado,
        )

        col_codigo = buscar_columna(
            encabezados,
            [
                "MATERIAL",
            ],
        )

        col_descripcion = buscar_columna(
            encabezados,
            [
                "TEXTO BREVE DE MATERIAL",
            ],
        )

        col_unidad = buscar_columna(
            encabezados,
            [
                "UNIDAD MEDIDA BASE",
                "UNIDAD DE MEDIDA BASE",
                "UM",
                "UNIDAD",
            ],
        )

        col_existencia = buscar_columna(
            encabezados,
            [
                "LIBRE UTILIZACION",
                "LIBRE UTIL.",
                "STOCK LIBRE UTILIZACION",
            ],
        )

        col_valor = buscar_columna(
            encabezados,
            [
                "VALOR LIBRE UTIL.",
                "VALOR LIBRE UTIL",
                "VALOR LIBRE UTILIZACION",
            ],
        )

        faltantes = []

        if not col_codigo:
            faltantes.append(
                "Material"
            )

        if not col_descripcion:
            faltantes.append(
                "Texto breve de material"
            )

        if not col_unidad:
            faltantes.append(
                "Unidad medida base"
            )

        if not col_existencia:
            faltantes.append(
                "Libre utilización"
            )

        if not col_valor:
            faltantes.append(
                "Valor libre util."
            )

        if faltantes:
            raise ValueError(
                "El archivo SAP no contiene las columnas requeridas: "
                + ", ".join(
                    faltantes
                )
            )

        fecha_carga = datetime.now().isoformat(
            timespec="seconds"
        )

        materiales = []
        snapshots = []

        cantidad = 0
        omitidos = 0

        indice_codigo = col_codigo - 1
        indice_descripcion = col_descripcion - 1
        indice_unidad = col_unidad - 1
        indice_existencia = col_existencia - 1
        indice_valor = col_valor - 1

        for fila in hoja.iter_rows(
            min_row=(
                fila_encabezado
                + 1
            ),
            values_only=True,
        ):
            codigo = normalizar_codigo_sap(
                fila[
                    indice_codigo
                ]
            )

            descripcion = str(
                fila[
                    indice_descripcion
                ]
                or ""
            ).strip()

            if not codigo:
                omitidos += 1
                continue

            if not descripcion:
                omitidos += 1
                continue

            unidad = str(
                fila[
                    indice_unidad
                ]
                or ""
            ).strip()

            existencia = convertir_numero(
                fila[
                    indice_existencia
                ]
            )

            valor_inventario = convertir_numero(
                fila[
                    indice_valor
                ]
            )

            valor_unitario = 0.0

            if existencia > 0:
                valor_unitario = (
                    valor_inventario
                    / existencia
                )

            materiales.append(
                (
                    codigo,
                    descripcion,
                    unidad,
                    existencia,
                    valor_inventario,
                    valor_unitario,
                    fecha_carga,
                )
            )

            snapshots.append(
                (
                    fecha_carga,
                    codigo,
                    descripcion,
                    unidad,
                    existencia,
                    valor_inventario,
                    valor_unitario,
                )
            )

            cantidad += 1

        conexion = conectar_bd()

        try:
            conexion.execute(
                """
                UPDATE materiales_sap
                SET activo = 0
                """
            )

            conexion.executemany(
                """
                INSERT INTO materiales_sap (
                    codigo_sap,
                    descripcion,
                    unidad_medida,
                    existencia,
                    valor_inventario,
                    valor_unitario,
                    origen_dato,
                    fecha_actualizacion,
                    activo
                )
                VALUES (
                    ?,
                    ?,
                    ?,
                    ?,
                    ?,
                    ?,
                    'SAP',
                    ?,
                    1
                )
                ON CONFLICT(codigo_sap)
                DO UPDATE SET
                    descripcion = excluded.descripcion,
                    unidad_medida = excluded.unidad_medida,
                    existencia = excluded.existencia,
                    valor_inventario = excluded.valor_inventario,
                    valor_unitario = excluded.valor_unitario,
                    fecha_actualizacion = excluded.fecha_actualizacion,
                    origen_dato = 'SAP',
                    activo = 1
                """,
                materiales,
            )

            conexion.executemany(
                """
                INSERT INTO snapshots_sap (
                    fecha_carga,
                    codigo_sap,
                    descripcion,
                    unidad_medida,
                    existencia,
                    valor_inventario,
                    valor_unitario
                )
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                snapshots,
            )

            conexion.execute(
                """
                INSERT INTO cargas_sap (
                    fecha_carga,
                    nombre_archivo,
                    cantidad_registros
                )
                VALUES (?, ?, ?)
                """,
                (
                    fecha_carga,
                    ruta_excel.name,
                    cantidad,
                ),
            )

            conexion.commit()

            return {
                "cantidad":
                    cantidad,

                "omitidos":
                    omitidos,

                "fecha_carga":
                    fecha_carga,

                "archivo":
                    ruta_excel.name,

                "fila_encabezado":
                    fila_encabezado,
            }

        except Exception:
            conexion.rollback()
            raise

        finally:
            conexion.close()

    finally:
        workbook.close()


def obtener_seguimiento_activo():
    conexion = conectar_bd()

    try:
        filas = conexion.execute(
            """
            SELECT
                sd.id AS detalle_id,
                sd.solicitud_id,
                sd.codigo_sap,
                sd.descripcion,
                sd.origen,
                sd.unidad_medida,
                sd.existencia_al_solicitar,
                sd.cantidad_pedir,
                sd.cantidad_recibida,
                sd.cantidad_retirada,
                sd.valor_unitario,
                sd.observaciones,
                sd.estado,
                s.codigo_solicitud,
                s.semana,
                s.anio,
                s.fecha_creacion,
                COALESCE(ms.existencia, 0) AS existencia_actual
            FROM solicitud_detalle sd
            INNER JOIN solicitudes s
                ON s.id = sd.solicitud_id
            LEFT JOIN materiales_sap ms
                ON ms.codigo_sap = sd.codigo_sap
            WHERE s.estado NOT IN ('CERRADA', 'ANULADA')
            AND sd.estado IN (
                'PENDIENTE',
                'RECIBIDO_PARCIAL',
                'PENDIENTE_RETIRO'
            )
            ORDER BY
                s.fecha_creacion DESC,
                sd.id DESC
            """
        ).fetchall()

        resultado = []

        for fila in filas:
            item = dict(fila)

            cantidad_pedir = float(
                item.get("cantidad_pedir") or 0
            )

            cantidad_recibida = float(
                item.get("cantidad_recibida") or 0
            )

            existencia_inicial = float(
                item.get("existencia_al_solicitar") or 0
            )

            existencia_actual = float(
                item.get("existencia_actual") or 0
            )

            pendiente = max(
                0,
                cantidad_pedir
                - cantidad_recibida,
            )

            aumento_stock = max(
                0,
                existencia_actual
                - existencia_inicial,
            )

            posible_recepcion = min(
                pendiente,
                aumento_stock,
            )

            item["cantidad_pendiente"] = pendiente
            item["aumento_stock"] = aumento_stock
            item["posible_recepcion"] = posible_recepcion

            if pendiente <= 0:
                estado_visual = "RECIBIDO"
            elif cantidad_recibida > 0:
                estado_visual = "RECIBIDO_PARCIAL"
            elif posible_recepcion > 0:
                estado_visual = "POSIBLE_RECEPCION"
            else:
                estado_visual = "PENDIENTE"

            item["estado_visual"] = estado_visual

            resultado.append(item)

        return resultado

    finally:
        conexion.close()


def confirmar_recepcion_detalle(
    detalle_id: int,
    cantidad: float,
):
    if cantidad <= 0:
        raise ValueError(
            "La cantidad recibida debe ser mayor que cero."
        )

    conexion = conectar_bd()

    try:
        fila = conexion.execute(
            """
            SELECT
                id,
                cantidad_pedir,
                cantidad_recibida,
                estado
            FROM solicitud_detalle
            WHERE id = ?
            """,
            (detalle_id,),
        ).fetchone()

        if fila is None:
            raise ValueError(
                "No se encontró el detalle solicitado."
            )

        cantidad_pedir = float(
            fila["cantidad_pedir"] or 0
        )

        cantidad_recibida_actual = float(
            fila["cantidad_recibida"] or 0
        )

        pendiente = max(
            0,
            cantidad_pedir
            - cantidad_recibida_actual,
        )

        if pendiente <= 0:
            raise ValueError(
                "Este material ya está completamente recibido."
            )

        if cantidad > pendiente:
            raise ValueError(
                f"La cantidad recibida no puede superar "
                f"la cantidad pendiente ({pendiente:g})."
            )

        nueva_cantidad_recibida = (
            cantidad_recibida_actual
            + cantidad
        )

        if nueva_cantidad_recibida >= cantidad_pedir:
            nuevo_estado = "PENDIENTE_RETIRO"
        else:
            nuevo_estado = "RECIBIDO_PARCIAL"

        fecha_recepcion = datetime.now().isoformat(
            timespec="seconds"
        )

        conexion.execute(
            """
            UPDATE solicitud_detalle
            SET
                cantidad_recibida = ?,
                estado = ?,
                fecha_recibido = ?
            WHERE id = ?
            """,
            (
                nueva_cantidad_recibida,
                nuevo_estado,
                fecha_recepcion,
                detalle_id,
            ),
        )

        conexion.execute(
            """
            INSERT INTO recepciones (
                detalle_solicitud_id,
                cantidad,
                fecha_recepcion,
                confirmado,
                observaciones
            )
            VALUES (?, ?, ?, 1, ?)
            """,
            (
                detalle_id,
                cantidad,
                fecha_recepcion,
                "Recepción confirmada por el usuario.",
            ),
        )

        conexion.commit()

        return {
            "detalle_id": detalle_id,
            "cantidad_confirmada": cantidad,
            "cantidad_recibida_total":
                nueva_cantidad_recibida,
            "cantidad_pendiente":
                max(
                    0,
                    cantidad_pedir
                    - nueva_cantidad_recibida,
                ),
            "estado": nuevo_estado,
        }

    except Exception:
        conexion.rollback()
        raise

    finally:
        conexion.close()


@repuestos_bp.route("/")
def inicio():
    return redirect(
        url_for(
            "repuestos.nueva_solicitud"
        )
    )


@repuestos_bp.route("/nueva")
def nueva_solicitud():
    inicializar_bd()

    return render_template(
        "repuestos/nueva_solicitud.html"
    )


@repuestos_bp.route(
    "/actualizar-sap",
    methods=[
        "GET",
        "POST",
    ],
)
def actualizar_sap():
    inicializar_bd()

    if request.method == "POST":
        archivo = request.files.get(
            "archivo_sap"
        )

        if (
            archivo is None
            or not archivo.filename
        ):
            flash(
                "Seleccione un archivo Excel.",
                "error",
            )

            return redirect(
                url_for(
                    "repuestos.actualizar_sap"
                )
            )

        extension = Path(
            archivo.filename
        ).suffix.lower()

        if extension not in (
            ".xlsx",
            ".xlsm",
        ):
            flash(
                "Solo se permiten archivos .xlsx o .xlsm.",
                "error",
            )

            return redirect(
                url_for(
                    "repuestos.actualizar_sap"
                )
            )

        carpeta = (
            CARPETA_DATOS
            / "cargas_sap"
        )

        carpeta.mkdir(
            parents=True,
            exist_ok=True,
        )

        nombre_original = Path(
            archivo.filename
        ).name

        nombre = (
            datetime.now().strftime(
                "%Y%m%d_%H%M%S"
            )
            + "_"
            + nombre_original
        )

        ruta = carpeta / nombre

        archivo.save(
            ruta
        )

        try:
            resultado = importar_base_sap(
                ruta
            )

            mensaje = (
                f"Base SAP actualizada correctamente. "
                f"{resultado['cantidad']} materiales cargados."
            )

            if resultado["omitidos"]:
                mensaje += (
                    f" {resultado['omitidos']} filas fueron omitidas "
                    f"por no tener código o descripción."
                )

            flash(
                mensaje,
                "exito",
            )

        except Exception as error:
            flash(
                f"No fue posible actualizar la base SAP: {error}",
                "error",
            )

        return redirect(
            url_for(
                "repuestos.actualizar_sap"
            )
        )

    conexion = conectar_bd()

    try:
        ultima_carga = conexion.execute(
            """
            SELECT
                fecha_carga,
                nombre_archivo,
                cantidad_registros
            FROM cargas_sap
            ORDER BY id DESC
            LIMIT 1
            """
        ).fetchone()

        total_materiales = conexion.execute(
            """
            SELECT COUNT(*) AS total
            FROM materiales_sap
            WHERE activo = 1
            """
        ).fetchone()["total"]

    finally:
        conexion.close()

    return render_template(
        "repuestos/actualizar_sap.html",
        ultima_carga=ultima_carga,
        total_materiales=int(
            total_materiales
            or 0
        ),
    )


def obtener_recibidos_por_retirar():
    conexion = conectar_bd()

    try:
        filas = conexion.execute(
            """
            SELECT
                sd.id AS detalle_id,
                sd.solicitud_id,
                sd.codigo_sap,
                sd.descripcion,
                sd.unidad_medida,
                sd.cantidad_pedir,
                sd.cantidad_recibida,
                sd.cantidad_retirada,
                sd.fecha_recibido,
                sd.fecha_retiro,
                sd.estado,
                sd.observaciones,
                s.codigo_solicitud,
                s.semana,
                s.anio,
                s.fecha_creacion
            FROM solicitud_detalle sd
            INNER JOIN solicitudes s
                ON s.id = sd.solicitud_id
            WHERE sd.cantidad_recibida > sd.cantidad_retirada
            AND s.estado <> 'ANULADA'
            ORDER BY
                COALESCE(sd.fecha_recibido, s.fecha_creacion) DESC,
                sd.id DESC
            """
        ).fetchall()

        resultado = []

        for fila in filas:
            item = dict(fila)

            cantidad_recibida = float(
                item.get(
                    "cantidad_recibida"
                )
                or 0
            )

            cantidad_retirada = float(
                item.get(
                    "cantidad_retirada"
                )
                or 0
            )

            item[
                "cantidad_pendiente_retiro"
            ] = max(
                0,
                cantidad_recibida
                - cantidad_retirada,
            )

            resultado.append(
                item
            )

        return resultado

    finally:
        conexion.close()


def confirmar_retiro_detalle(
    detalle_id: int,
    cantidad: float,
    observaciones: str = "",
):
    if cantidad <= 0:
        raise ValueError(
            "La cantidad retirada debe ser mayor que cero."
        )

    conexion = conectar_bd()

    try:
        fila = conexion.execute(
            """
            SELECT
                id,
                solicitud_id,
                cantidad_pedir,
                cantidad_recibida,
                cantidad_retirada,
                estado
            FROM solicitud_detalle
            WHERE id = ?
            """,
            (
                detalle_id,
            ),
        ).fetchone()

        if fila is None:
            raise ValueError(
                "No se encontró el material seleccionado."
            )

        cantidad_pedir = float(
            fila[
                "cantidad_pedir"
            ]
            or 0
        )

        cantidad_recibida = float(
            fila[
                "cantidad_recibida"
            ]
            or 0
        )

        cantidad_retirada_actual = float(
            fila[
                "cantidad_retirada"
            ]
            or 0
        )

        pendiente_retiro = max(
            0,
            cantidad_recibida
            - cantidad_retirada_actual,
        )

        if pendiente_retiro <= 0:
            raise ValueError(
                "Este material no tiene unidades pendientes por retirar."
            )

        if cantidad > pendiente_retiro:
            raise ValueError(
                f"La cantidad retirada no puede superar "
                f"la cantidad pendiente ({pendiente_retiro:g})."
            )

        nueva_cantidad_retirada = (
            cantidad_retirada_actual
            + cantidad
        )

        pendiente_recepcion = max(
            0,
            cantidad_pedir
            - cantidad_recibida,
        )

        pendiente_retiro_nuevo = max(
            0,
            cantidad_recibida
            - nueva_cantidad_retirada,
        )

        if (
            pendiente_recepcion <= 0
            and pendiente_retiro_nuevo <= 0
        ):
            nuevo_estado = "COMPLETADO"

        elif pendiente_recepcion > 0:
            nuevo_estado = "RECIBIDO_PARCIAL"

        else:
            nuevo_estado = "PENDIENTE_RETIRO"

        fecha_retiro = datetime.now().isoformat(
            timespec="seconds"
        )

        conexion.execute(
            """
            UPDATE solicitud_detalle
            SET
                cantidad_retirada = ?,
                estado = ?,
                fecha_retiro = ?
            WHERE id = ?
            """,
            (
                nueva_cantidad_retirada,
                nuevo_estado,
                fecha_retiro,
                detalle_id,
            ),
        )

        conexion.execute(
            """
            INSERT INTO retiros (
                detalle_solicitud_id,
                cantidad,
                fecha_retiro,
                observaciones
            )
            VALUES (
                ?,
                ?,
                ?,
                ?
            )
            """,
            (
                detalle_id,
                cantidad,
                fecha_retiro,
                observaciones.strip(),
            ),
        )

        solicitud_id = int(
            fila[
                "solicitud_id"
            ]
        )

        pendiente_solicitud = conexion.execute(
            """
            SELECT COUNT(*) AS total
            FROM solicitud_detalle
            WHERE solicitud_id = ?
            AND estado <> 'COMPLETADO'
            """,
            (
                solicitud_id,
            ),
        ).fetchone()[
            "total"
        ]

        if int(
            pendiente_solicitud
            or 0
        ) == 0:
            conexion.execute(
                """
                UPDATE solicitudes
                SET estado = 'CERRADA'
                WHERE id = ?
                """,
                (
                    solicitud_id,
                ),
            )

        conexion.commit()

        return {
            "detalle_id":
                detalle_id,

            "cantidad_retirada":
                cantidad,

            "cantidad_retirada_total":
                nueva_cantidad_retirada,

            "cantidad_pendiente_retiro":
                pendiente_retiro_nuevo,

            "cantidad_pendiente_recepcion":
                pendiente_recepcion,

            "estado":
                nuevo_estado,
        }

    except Exception:
        conexion.rollback()
        raise

    finally:
        conexion.close()


def generar_excel_retiros():
    load_workbook = cargar_openpyxl()

    from openpyxl import Workbook

    registros = obtener_recibidos_por_retirar()

    if not registros:
        raise ValueError(
            "No hay materiales pendientes por retirar."
        )

    carpeta = (
        BASE_DIR
        / "salidas"
        / "repuestos"
        / "retiros"
    )

    carpeta.mkdir(
        parents=True,
        exist_ok=True,
    )

    fecha = datetime.now()

    nombre = (
        "MATERIALES_PENDIENTES_RETIRO_"
        + fecha.strftime(
            "%Y%m%d_%H%M%S"
        )
        + ".xlsx"
    )

    ruta = carpeta / nombre

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Pendientes por retirar"

    encabezados = [
        "Solicitud",
        "Semana",
        "Código SAP",
        "Descripción",
        "Unidad",
        "Cantidad recibida",
        "Cantidad retirada",
        "Pendiente por retirar",
        "Fecha recibido",
        "Observaciones",
    ]

    hoja.append(
        encabezados
    )

    for item in registros:
        hoja.append(
            [
                item.get(
                    "codigo_solicitud",
                    "",
                ),
                item.get(
                    "semana",
                    "",
                ),
                item.get(
                    "codigo_sap",
                    "",
                )
                or "CREAR",
                item.get(
                    "descripcion",
                    "",
                ),
                item.get(
                    "unidad_medida",
                    "",
                ),
                item.get(
                    "cantidad_recibida",
                    0,
                ),
                item.get(
                    "cantidad_retirada",
                    0,
                ),
                item.get(
                    "cantidad_pendiente_retiro",
                    0,
                ),
                item.get(
                    "fecha_recibido",
                    "",
                ),
                item.get(
                    "observaciones",
                    "",
                ),
            ]
        )

    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions

    anchos = {
        "A": 20,
        "B": 10,
        "C": 18,
        "D": 50,
        "E": 12,
        "F": 18,
        "G": 18,
        "H": 22,
        "I": 22,
        "J": 45,
    }

    for columna, ancho in anchos.items():
        hoja.column_dimensions[
            columna
        ].width = ancho

    libro.save(
        ruta
    )

    return ruta


def obtener_secciones_inventario(
    incluir_inactivas: bool = True,
):
    conexion = conectar_bd()

    try:
        condicion = ""

        if not incluir_inactivas:
            condicion = "WHERE s.activo = 1"

        filas = conexion.execute(
            f"""
            SELECT
                s.id,
                s.nombre,
                s.descripcion,
                s.activo,
                s.fecha_creacion,
                COUNT(DISTINCT ss.id) AS total_subsecciones,
                COUNT(DISTINCT it.id) AS total_materiales
            FROM secciones s
            LEFT JOIN subsecciones ss
                ON ss.seccion_id = s.id
            LEFT JOIN inventario_tecnico it
                ON it.seccion_id = s.id
                AND it.activo = 1
            {condicion}
            GROUP BY
                s.id,
                s.nombre,
                s.descripcion,
                s.activo,
                s.fecha_creacion
            ORDER BY
                s.activo DESC,
                UPPER(s.nombre)
            """
        ).fetchall()

        return [
            dict(
                fila
            )
            for fila in filas
        ]

    finally:
        conexion.close()


def obtener_subsecciones_inventario(
    incluir_inactivas: bool = True,
):
    conexion = conectar_bd()

    try:
        condicion = ""

        if not incluir_inactivas:
            condicion = "WHERE ss.activo = 1"

        filas = conexion.execute(
            f"""
            SELECT
                ss.id,
                ss.seccion_id,
                ss.nombre,
                ss.descripcion,
                ss.activo,
                ss.fecha_creacion,
                s.nombre AS seccion_nombre,
                s.activo AS seccion_activa,
                COUNT(DISTINCT it.id) AS total_materiales
            FROM subsecciones ss
            INNER JOIN secciones s
                ON s.id = ss.seccion_id
            LEFT JOIN inventario_tecnico it
                ON it.subseccion_id = ss.id
                AND it.activo = 1
            {condicion}
            GROUP BY
                ss.id,
                ss.seccion_id,
                ss.nombre,
                ss.descripcion,
                ss.activo,
                ss.fecha_creacion,
                s.nombre,
                s.activo
            ORDER BY
                s.nombre,
                ss.activo DESC,
                UPPER(ss.nombre)
            """
        ).fetchall()

        return [
            dict(
                fila
            )
            for fila in filas
        ]

    finally:
        conexion.close()


def obtener_pedidos_pendientes_por_codigo(
    codigo_sap: str,
) -> float:
    if not codigo_sap:
        return 0.0

    conexion = conectar_bd()

    try:
        fila = conexion.execute(
            """
            SELECT
                COALESCE(
                    SUM(
                        MAX(
                            cantidad_pedir
                            - cantidad_recibida,
                            0
                        )
                    ),
                    0
                ) AS pendiente
            FROM solicitud_detalle
            WHERE codigo_sap = ?
            AND estado IN (
                'PENDIENTE',
                'RECIBIDO_PARCIAL'
            )
            """,
            (
                codigo_sap,
            ),
        ).fetchone()

        return float(
            fila[
                "pendiente"
            ]
            or 0
        )

    finally:
        conexion.close()


def obtener_inventario_tecnico():
    conexion = conectar_bd()

    try:
        filas = conexion.execute(
            """
            SELECT
                it.id,
                it.codigo_sap,
                it.descripcion,
                it.unidad_medida,
                it.seccion_id,
                it.subseccion_id,
                it.tipo_control,
                it.stock_minimo,
                it.stock_objetivo,
                it.origen_dato,
                it.observaciones,
                it.activo,
                s.nombre AS seccion_nombre,
                s.activo AS seccion_activa,
                ss.nombre AS subseccion_nombre,
                ss.activo AS subseccion_activa,
                COALESCE(ms.existencia, 0) AS stock_sap,
                COALESCE(ms.valor_unitario, 0) AS valor_unitario
            FROM inventario_tecnico it
            INNER JOIN secciones s
                ON s.id = it.seccion_id
            LEFT JOIN subsecciones ss
                ON ss.id = it.subseccion_id
            LEFT JOIN materiales_sap ms
                ON ms.codigo_sap = it.codigo_sap
            ORDER BY
                it.activo DESC,
                s.nombre,
                COALESCE(ss.nombre, ''),
                it.descripcion
            """
        ).fetchall()

        resultado = []

        for fila in filas:
            item = dict(
                fila
            )

            codigo = str(
                item.get(
                    "codigo_sap",
                    ""
                )
                or ""
            ).strip()

            pendiente = 0.0

            if codigo and codigo != "CREAR":
                pendiente = (
                    obtener_pedidos_pendientes_por_codigo(
                        codigo
                    )
                )

            stock_sap = float(
                item.get(
                    "stock_sap"
                )
                or 0
            )

            stock_minimo = float(
                item.get(
                    "stock_minimo"
                )
                or 0
            )

            stock_objetivo = float(
                item.get(
                    "stock_objetivo"
                )
                or 0
            )

            stock_proyectado = (
                stock_sap
                + pendiente
            )

            faltante_minimo = max(
                0,
                stock_minimo
                - stock_proyectado,
            )

            faltante_objetivo = max(
                0,
                stock_objetivo
                - stock_proyectado,
            )

            if not int(
                item.get(
                    "activo"
                )
                or 0
            ):
                estado_stock = "INACTIVO"

            elif stock_proyectado < stock_minimo:
                estado_stock = "CRITICO"

            elif (
                stock_objetivo > 0
                and stock_proyectado
                < stock_objetivo
            ):
                estado_stock = "BAJO"

            else:
                estado_stock = "OK"

            item[
                "pedido_pendiente"
            ] = pendiente

            item[
                "stock_proyectado"
            ] = stock_proyectado

            item[
                "faltante_minimo"
            ] = faltante_minimo

            item[
                "faltante_objetivo"
            ] = faltante_objetivo

            item[
                "estado_stock"
            ] = estado_stock

            resultado.append(
                item
            )

        return resultado

    finally:
        conexion.close()


def crear_seccion(
    nombre: str,
    descripcion: str = "",
):
    nombre = str(
        nombre
        or ""
    ).strip()

    descripcion = str(
        descripcion
        or ""
    ).strip()

    if not nombre:
        raise ValueError(
            "Ingrese el nombre de la sección."
        )

    conexion = conectar_bd()

    try:
        existente = conexion.execute(
            """
            SELECT id
            FROM secciones
            WHERE UPPER(nombre) = UPPER(?)
            """,
            (
                nombre,
            ),
        ).fetchone()

        if existente:
            raise ValueError(
                "Ya existe una sección con ese nombre."
            )

        conexion.execute(
            """
            INSERT INTO secciones (
                nombre,
                descripcion,
                activo,
                fecha_creacion
            )
            VALUES (
                ?,
                ?,
                1,
                ?
            )
            """,
            (
                nombre,
                descripcion,
                datetime.now().isoformat(
                    timespec="seconds"
                ),
            ),
        )

        conexion.commit()

    except Exception:
        conexion.rollback()
        raise

    finally:
        conexion.close()


def editar_seccion(
    seccion_id: int,
    nombre: str,
    descripcion: str = "",
):
    nombre = str(
        nombre
        or ""
    ).strip()

    descripcion = str(
        descripcion
        or ""
    ).strip()

    if not nombre:
        raise ValueError(
            "Ingrese el nombre de la sección."
        )

    conexion = conectar_bd()

    try:
        duplicada = conexion.execute(
            """
            SELECT id
            FROM secciones
            WHERE UPPER(nombre) = UPPER(?)
            AND id <> ?
            """,
            (
                nombre,
                seccion_id,
            ),
        ).fetchone()

        if duplicada:
            raise ValueError(
                "Ya existe otra sección con ese nombre."
            )

        cursor = conexion.execute(
            """
            UPDATE secciones
            SET
                nombre = ?,
                descripcion = ?
            WHERE id = ?
            """,
            (
                nombre,
                descripcion,
                seccion_id,
            ),
        )

        if cursor.rowcount == 0:
            raise ValueError(
                "La sección no existe."
            )

        conexion.commit()

    except Exception:
        conexion.rollback()
        raise

    finally:
        conexion.close()


def cambiar_estado_seccion(
    seccion_id: int,
    activo: bool,
):
    conexion = conectar_bd()

    try:
        cursor = conexion.execute(
            """
            UPDATE secciones
            SET activo = ?
            WHERE id = ?
            """,
            (
                1 if activo else 0,
                seccion_id,
            ),
        )

        if cursor.rowcount == 0:
            raise ValueError(
                "La sección no existe."
            )

        if not activo:
            conexion.execute(
                """
                UPDATE subsecciones
                SET activo = 0
                WHERE seccion_id = ?
                """,
                (
                    seccion_id,
                ),
            )

        conexion.commit()

    except Exception:
        conexion.rollback()
        raise

    finally:
        conexion.close()


def eliminar_seccion_definitiva(
    seccion_id: int,
):
    conexion = conectar_bd()

    try:
        total_subsecciones = conexion.execute(
            """
            SELECT COUNT(*) AS total
            FROM subsecciones
            WHERE seccion_id = ?
            """,
            (
                seccion_id,
            ),
        ).fetchone()[
            "total"
        ]

        total_materiales = conexion.execute(
            """
            SELECT COUNT(*) AS total
            FROM inventario_tecnico
            WHERE seccion_id = ?
            """,
            (
                seccion_id,
            ),
        ).fetchone()[
            "total"
        ]

        if int(
            total_subsecciones
            or 0
        ) > 0:
            raise ValueError(
                "No se puede eliminar definitivamente la sección "
                "porque todavía tiene subsecciones."
            )

        if int(
            total_materiales
            or 0
        ) > 0:
            raise ValueError(
                "No se puede eliminar definitivamente la sección "
                "porque todavía tiene materiales asociados."
            )

        cursor = conexion.execute(
            """
            DELETE FROM secciones
            WHERE id = ?
            """,
            (
                seccion_id,
            ),
        )

        if cursor.rowcount == 0:
            raise ValueError(
                "La sección no existe."
            )

        conexion.commit()

    except Exception:
        conexion.rollback()
        raise

    finally:
        conexion.close()


def crear_subseccion(
    seccion_id: int,
    nombre: str,
    descripcion: str = "",
):
    nombre = str(
        nombre
        or ""
    ).strip()

    descripcion = str(
        descripcion
        or ""
    ).strip()

    if not nombre:
        raise ValueError(
            "Ingrese el nombre de la subsección."
        )

    conexion = conectar_bd()

    try:
        seccion = conexion.execute(
            """
            SELECT
                id,
                activo
            FROM secciones
            WHERE id = ?
            """,
            (
                seccion_id,
            ),
        ).fetchone()

        if seccion is None:
            raise ValueError(
                "La sección seleccionada no existe."
            )

        if not int(
            seccion[
                "activo"
            ]
            or 0
        ):
            raise ValueError(
                "La sección está inactiva."
            )

        duplicada = conexion.execute(
            """
            SELECT id
            FROM subsecciones
            WHERE seccion_id = ?
            AND UPPER(nombre) = UPPER(?)
            """,
            (
                seccion_id,
                nombre,
            ),
        ).fetchone()

        if duplicada:
            raise ValueError(
                "Ya existe una subsección con ese nombre "
                "dentro de la sección."
            )

        conexion.execute(
            """
            INSERT INTO subsecciones (
                seccion_id,
                nombre,
                descripcion,
                activo,
                fecha_creacion
            )
            VALUES (
                ?,
                ?,
                ?,
                1,
                ?
            )
            """,
            (
                seccion_id,
                nombre,
                descripcion,
                datetime.now().isoformat(
                    timespec="seconds"
                ),
            ),
        )

        conexion.commit()

    except Exception:
        conexion.rollback()
        raise

    finally:
        conexion.close()


def editar_subseccion(
    subseccion_id: int,
    nombre: str,
    descripcion: str = "",
):
    nombre = str(
        nombre
        or ""
    ).strip()

    descripcion = str(
        descripcion
        or ""
    ).strip()

    if not nombre:
        raise ValueError(
            "Ingrese el nombre de la subsección."
        )

    conexion = conectar_bd()

    try:
        actual = conexion.execute(
            """
            SELECT seccion_id
            FROM subsecciones
            WHERE id = ?
            """,
            (
                subseccion_id,
            ),
        ).fetchone()

        if actual is None:
            raise ValueError(
                "La subsección no existe."
            )

        duplicada = conexion.execute(
            """
            SELECT id
            FROM subsecciones
            WHERE seccion_id = ?
            AND UPPER(nombre) = UPPER(?)
            AND id <> ?
            """,
            (
                actual[
                    "seccion_id"
                ],
                nombre,
                subseccion_id,
            ),
        ).fetchone()

        if duplicada:
            raise ValueError(
                "Ya existe otra subsección con ese nombre."
            )

        conexion.execute(
            """
            UPDATE subsecciones
            SET
                nombre = ?,
                descripcion = ?
            WHERE id = ?
            """,
            (
                nombre,
                descripcion,
                subseccion_id,
            ),
        )

        conexion.commit()

    except Exception:
        conexion.rollback()
        raise

    finally:
        conexion.close()


def cambiar_estado_subseccion(
    subseccion_id: int,
    activo: bool,
):
    conexion = conectar_bd()

    try:
        if activo:
            fila = conexion.execute(
                """
                SELECT
                    s.activo AS seccion_activa
                FROM subsecciones ss
                INNER JOIN secciones s
                    ON s.id = ss.seccion_id
                WHERE ss.id = ?
                """,
                (
                    subseccion_id,
                ),
            ).fetchone()

            if fila is None:
                raise ValueError(
                    "La subsección no existe."
                )

            if not int(
                fila[
                    "seccion_activa"
                ]
                or 0
            ):
                raise ValueError(
                    "No puede reactivar esta subsección "
                    "mientras su sección esté inactiva."
                )

        cursor = conexion.execute(
            """
            UPDATE subsecciones
            SET activo = ?
            WHERE id = ?
            """,
            (
                1 if activo else 0,
                subseccion_id,
            ),
        )

        if cursor.rowcount == 0:
            raise ValueError(
                "La subsección no existe."
            )

        conexion.commit()

    except Exception:
        conexion.rollback()
        raise

    finally:
        conexion.close()


def eliminar_subseccion_definitiva(
    subseccion_id: int,
):
    conexion = conectar_bd()

    try:
        total_materiales = conexion.execute(
            """
            SELECT COUNT(*) AS total
            FROM inventario_tecnico
            WHERE subseccion_id = ?
            """,
            (
                subseccion_id,
            ),
        ).fetchone()[
            "total"
        ]

        if int(
            total_materiales
            or 0
        ) > 0:
            raise ValueError(
                "No se puede eliminar definitivamente la subsección "
                "porque todavía tiene materiales asociados."
            )

        cursor = conexion.execute(
            """
            DELETE FROM subsecciones
            WHERE id = ?
            """,
            (
                subseccion_id,
            ),
        )

        if cursor.rowcount == 0:
            raise ValueError(
                "La subsección no existe."
            )

        conexion.commit()

    except Exception:
        conexion.rollback()
        raise

    finally:
        conexion.close()


def crear_material_inventario(
    codigo_sap: str,
    descripcion: str,
    unidad_medida: str,
    seccion_id: int,
    subseccion_id: int | None,
    stock_minimo: float,
    stock_objetivo: float,
    observaciones: str = "",
    origen_dato: str = "SAP",
):
    codigo_sap = str(
        codigo_sap
        or ""
    ).strip()

    descripcion = str(
        descripcion
        or ""
    ).strip()

    unidad_medida = str(
        unidad_medida
        or ""
    ).strip()

    observaciones = str(
        observaciones
        or ""
    ).strip()

    origen_dato = str(
        origen_dato
        or "SAP"
    ).strip().upper()

    if not descripcion:
        raise ValueError(
            "Ingrese la descripción del material."
        )

    if not unidad_medida:
        raise ValueError(
            "Ingrese la unidad de medida."
        )

    if stock_minimo < 0:
        raise ValueError(
            "El stock mínimo no puede ser negativo."
        )

    if stock_objetivo < stock_minimo:
        raise ValueError(
            "El stock objetivo debe ser igual o mayor "
            "que el stock mínimo."
        )

    conexion = conectar_bd()

    try:
        seccion = conexion.execute(
            """
            SELECT activo
            FROM secciones
            WHERE id = ?
            """,
            (
                seccion_id,
            ),
        ).fetchone()

        if seccion is None:
            raise ValueError(
                "La sección seleccionada no existe."
            )

        if not int(
            seccion[
                "activo"
            ]
            or 0
        ):
            raise ValueError(
                "La sección seleccionada está inactiva."
            )

        if subseccion_id:
            subseccion = conexion.execute(
                """
                SELECT
                    seccion_id,
                    activo
                FROM subsecciones
                WHERE id = ?
                """,
                (
                    subseccion_id,
                ),
            ).fetchone()

            if subseccion is None:
                raise ValueError(
                    "La subsección seleccionada no existe."
                )

            if int(
                subseccion[
                    "seccion_id"
                ]
            ) != int(
                seccion_id
            ):
                raise ValueError(
                    "La subsección no pertenece a la sección seleccionada."
                )

            if not int(
                subseccion[
                    "activo"
                ]
                or 0
            ):
                raise ValueError(
                    "La subsección seleccionada está inactiva."
                )

        if codigo_sap and codigo_sap != "CREAR":
            duplicado = conexion.execute(
                """
                SELECT id
                FROM inventario_tecnico
                WHERE codigo_sap = ?
                AND seccion_id = ?
                AND COALESCE(subseccion_id, 0)
                    = COALESCE(?, 0)
                AND activo = 1
                """,
                (
                    codigo_sap,
                    seccion_id,
                    subseccion_id,
                ),
            ).fetchone()

            if duplicado:
                raise ValueError(
                    "Este material ya está agregado en esa "
                    "sección/subsección."
                )

        conexion.execute(
            """
            INSERT INTO inventario_tecnico (
                codigo_sap,
                descripcion,
                unidad_medida,
                seccion_id,
                subseccion_id,
                tipo_control,
                stock_minimo,
                stock_objetivo,
                factor_minimo,
                factor_objetivo,
                origen_dato,
                observaciones,
                activo,
                fecha_creacion
            )
            VALUES (
                ?,
                ?,
                ?,
                ?,
                ?,
                'stock',
                ?,
                ?,
                1,
                1,
                ?,
                ?,
                1,
                ?
            )
            """,
            (
                codigo_sap
                or "CREAR",
                descripcion,
                unidad_medida,
                seccion_id,
                subseccion_id,
                stock_minimo,
                stock_objetivo,
                origen_dato,
                observaciones,
                datetime.now().isoformat(
                    timespec="seconds"
                ),
            ),
        )

        conexion.commit()

    except Exception:
        conexion.rollback()
        raise

    finally:
        conexion.close()


def editar_material_inventario(
    inventario_id: int,
    seccion_id: int,
    subseccion_id: int | None,
    stock_minimo: float,
    stock_objetivo: float,
    observaciones: str = "",
):
    if stock_minimo < 0:
        raise ValueError(
            "El stock mínimo no puede ser negativo."
        )

    if stock_objetivo < stock_minimo:
        raise ValueError(
            "El stock objetivo debe ser igual o mayor "
            "que el stock mínimo."
        )

    conexion = conectar_bd()

    try:
        seccion = conexion.execute(
            """
            SELECT activo
            FROM secciones
            WHERE id = ?
            """,
            (
                seccion_id,
            ),
        ).fetchone()

        if seccion is None:
            raise ValueError(
                "La sección no existe."
            )

        if subseccion_id:
            subseccion = conexion.execute(
                """
                SELECT
                    seccion_id,
                    activo
                FROM subsecciones
                WHERE id = ?
                """,
                (
                    subseccion_id,
                ),
            ).fetchone()

            if subseccion is None:
                raise ValueError(
                    "La subsección no existe."
                )

            if int(
                subseccion[
                    "seccion_id"
                ]
            ) != int(
                seccion_id
            ):
                raise ValueError(
                    "La subsección no pertenece a la sección."
                )

        cursor = conexion.execute(
            """
            UPDATE inventario_tecnico
            SET
                seccion_id = ?,
                subseccion_id = ?,
                stock_minimo = ?,
                stock_objetivo = ?,
                observaciones = ?
            WHERE id = ?
            """,
            (
                seccion_id,
                subseccion_id,
                stock_minimo,
                stock_objetivo,
                str(
                    observaciones
                    or ""
                ).strip(),
                inventario_id,
            ),
        )

        if cursor.rowcount == 0:
            raise ValueError(
                "El material de inventario no existe."
            )

        conexion.commit()

    except Exception:
        conexion.rollback()
        raise

    finally:
        conexion.close()


def cambiar_estado_material_inventario(
    inventario_id: int,
    activo: bool,
):
    conexion = conectar_bd()

    try:
        cursor = conexion.execute(
            """
            UPDATE inventario_tecnico
            SET activo = ?
            WHERE id = ?
            """,
            (
                1 if activo else 0,
                inventario_id,
            ),
        )

        if cursor.rowcount == 0:
            raise ValueError(
                "El material de inventario no existe."
            )

        conexion.commit()

    except Exception:
        conexion.rollback()
        raise

    finally:
        conexion.close()


def eliminar_material_inventario(
    inventario_id: int,
):
    conexion = conectar_bd()

    try:
        total_maquinas = conexion.execute(
            """
            SELECT COUNT(*) AS total
            FROM inventario_maquinas
            WHERE inventario_id = ?
            """,
            (
                inventario_id,
            ),
        ).fetchone()[
            "total"
        ]

        if int(
            total_maquinas
            or 0
        ) > 0:
            raise ValueError(
                "No se puede eliminar definitivamente este material "
                "porque tiene máquinas asociadas. Puede desactivarlo."
            )

        cursor = conexion.execute(
            """
            DELETE FROM inventario_tecnico
            WHERE id = ?
            """,
            (
                inventario_id,
            ),
        )

        if cursor.rowcount == 0:
            raise ValueError(
                "El material no existe."
            )

        conexion.commit()

    except Exception:
        conexion.rollback()
        raise

    finally:
        conexion.close()


@repuestos_bp.route("/seguimiento")
def seguimiento():
    inicializar_bd()

    registros = obtener_seguimiento_activo()

    resumen = {
        "total":
            len(registros),

        "pendientes":
            sum(
                1
                for item in registros
                if item["estado_visual"]
                == "PENDIENTE"
            ),

        "posibles":
            sum(
                1
                for item in registros
                if item["estado_visual"]
                == "POSIBLE_RECEPCION"
            ),

        "parciales":
            sum(
                1
                for item in registros
                if item["estado_visual"]
                == "RECIBIDO_PARCIAL"
            ),
    }

    return render_template(
        "repuestos/seguimiento.html",
        registros=registros,
        resumen=resumen,
    )

@repuestos_bp.route(
    "/api/inventario/catalogo"
)
def api_catalogo_inventario():
    inicializar_bd()

    try:
        secciones = obtener_secciones_inventario(
            incluir_inactivas=False
        )

        subsecciones = obtener_subsecciones_inventario(
            incluir_inactivas=False
        )

        return jsonify(
            {
                "ok": True,
                "secciones": secciones,
                "subsecciones": subsecciones,
            }
        )

    except Exception as error:
        return jsonify(
            {
                "ok": False,
                "error":
                    f"No fue posible consultar el inventario: {error}",
            }
        ), 500

@repuestos_bp.route(
    "/api/seguimiento/<int:detalle_id>/confirmar-recepcion",
    methods=["POST"],
)
def api_confirmar_recepcion(
    detalle_id,
):
    datos = request.get_json(
        silent=True
    ) or {}

    cantidad = convertir_numero(
        datos.get(
            "cantidad"
        )
    )

    try:
        resultado = confirmar_recepcion_detalle(
            detalle_id,
            cantidad,
        )

        return jsonify(
            {
                "ok": True,
                "resultado": resultado,
            }
        )

    except ValueError as error:
        return jsonify(
            {
                "ok": False,
                "error": str(error),
            }
        ), 400

    except Exception as error:
        return jsonify(
            {
                "ok": False,
                "error":
                    f"No fue posible confirmar la recepción: {error}",
            }
        ), 500


@repuestos_bp.route("/recibidos")
def recibidos_por_retirar():
    inicializar_bd()

    registros = obtener_recibidos_por_retirar()

    resumen = {
        "materiales":
            len(
                registros
            ),

        "unidades":
            sum(
                float(
                    item.get(
                        "cantidad_pendiente_retiro"
                    )
                    or 0
                )
                for item in registros
            ),
    }

    return render_template(
        "repuestos/recibidos.html",
        registros=registros,
        resumen=resumen,
    )


@repuestos_bp.route(
    "/api/recibidos/<int:detalle_id>/confirmar-retiro",
    methods=["POST"],
)
def api_confirmar_retiro(
    detalle_id,
):
    datos = request.get_json(
        silent=True
    ) or {}

    cantidad = convertir_numero(
        datos.get(
            "cantidad"
        )
    )

    observaciones = str(
        datos.get(
            "observaciones",
            ""
        )
    ).strip()

    try:
        resultado = confirmar_retiro_detalle(
            detalle_id,
            cantidad,
            observaciones,
        )

        return jsonify(
            {
                "ok":
                    True,

                "resultado":
                    resultado,
            }
        )

    except ValueError as error:
        return jsonify(
            {
                "ok":
                    False,

                "error":
                    str(
                        error
                    ),
            }
        ), 400

    except Exception as error:
        return jsonify(
            {
                "ok":
                    False,

                "error":
                    f"No fue posible confirmar el retiro: {error}",
            }
        ), 500


@repuestos_bp.route(
    "/recibidos/exportar"
)
def exportar_recibidos():
    try:
        ruta = generar_excel_retiros()

        return send_file(
            ruta,
            as_attachment=True,
            download_name=ruta.name,
        )

    except ValueError as error:
        flash(
            str(
                error
            ),
            "error",
        )

        return redirect(
            url_for(
                "repuestos.recibidos_por_retirar"
            )
        )


@repuestos_bp.route("/inventario")
def inventario_tecnico():
    inicializar_bd()

    secciones = obtener_secciones_inventario(
        incluir_inactivas=True
    )

    subsecciones = obtener_subsecciones_inventario(
        incluir_inactivas=True
    )

    inventario = obtener_inventario_tecnico()

    resumen = {
        "total":
            sum(
                1
                for item in inventario
                if int(
                    item.get(
                        "activo"
                    )
                    or 0
                )
            ),

        "criticos":
            sum(
                1
                for item in inventario
                if item.get(
                    "estado_stock"
                ) == "CRITICO"
            ),

        "bajos":
            sum(
                1
                for item in inventario
                if item.get(
                    "estado_stock"
                ) == "BAJO"
            ),

        "ok":
            sum(
                1
                for item in inventario
                if item.get(
                    "estado_stock"
                ) == "OK"
            ),
    }

    return render_template(
        "repuestos/inventario.html",
        secciones=secciones,
        subsecciones=subsecciones,
        inventario=inventario,
        resumen=resumen,
    )


@repuestos_bp.route(
    "/api/inventario/secciones",
    methods=["POST"],
)
def api_crear_seccion():
    datos = request.get_json(
        silent=True
    ) or {}

    try:
        crear_seccion(
            datos.get(
                "nombre",
                ""
            ),
            datos.get(
                "descripcion",
                ""
            ),
        )

        return jsonify(
            {
                "ok":
                    True
            }
        )

    except ValueError as error:
        return jsonify(
            {
                "ok":
                    False,
                "error":
                    str(
                        error
                    ),
            }
        ), 400

    except Exception as error:
        return jsonify(
            {
                "ok":
                    False,
                "error":
                    f"No fue posible crear la sección: {error}",
            }
        ), 500


@repuestos_bp.route(
    "/api/inventario/secciones/<int:seccion_id>",
    methods=["PUT", "DELETE"],
)
def api_gestionar_seccion(
    seccion_id,
):
    datos = request.get_json(
        silent=True
    ) or {}

    try:
        if request.method == "DELETE":
            eliminar_seccion_definitiva(
                seccion_id
            )

        else:
            accion = str(
                datos.get(
                    "accion",
                    "editar"
                )
            ).strip().lower()

            if accion == "editar":
                editar_seccion(
                    seccion_id,
                    datos.get(
                        "nombre",
                        ""
                    ),
                    datos.get(
                        "descripcion",
                        ""
                    ),
                )

            elif accion == "desactivar":
                cambiar_estado_seccion(
                    seccion_id,
                    False,
                )

            elif accion == "reactivar":
                cambiar_estado_seccion(
                    seccion_id,
                    True,
                )

            else:
                raise ValueError(
                    "Acción de sección no válida."
                )

        return jsonify(
            {
                "ok":
                    True
            }
        )

    except ValueError as error:
        return jsonify(
            {
                "ok":
                    False,
                "error":
                    str(
                        error
                    ),
            }
        ), 400

    except Exception as error:
        return jsonify(
            {
                "ok":
                    False,
                "error":
                    f"No fue posible actualizar la sección: {error}",
            }
        ), 500


@repuestos_bp.route(
    "/api/inventario/subsecciones",
    methods=["POST"],
)
def api_crear_subseccion():
    datos = request.get_json(
        silent=True
    ) or {}

    try:
        seccion_id = int(
            datos.get(
                "seccion_id"
            )
        )

        crear_subseccion(
            seccion_id,
            datos.get(
                "nombre",
                ""
            ),
            datos.get(
                "descripcion",
                ""
            ),
        )

        return jsonify(
            {
                "ok":
                    True
            }
        )

    except (
        TypeError,
        ValueError,
    ) as error:
        return jsonify(
            {
                "ok":
                    False,
                "error":
                    str(
                        error
                    )
                    or "Seleccione una sección.",
            }
        ), 400

    except Exception as error:
        return jsonify(
            {
                "ok":
                    False,
                "error":
                    f"No fue posible crear la subsección: {error}",
            }
        ), 500


@repuestos_bp.route(
    "/api/inventario/subsecciones/<int:subseccion_id>",
    methods=["PUT", "DELETE"],
)
def api_gestionar_subseccion(
    subseccion_id,
):
    datos = request.get_json(
        silent=True
    ) or {}

    try:
        if request.method == "DELETE":
            eliminar_subseccion_definitiva(
                subseccion_id
            )

        else:
            accion = str(
                datos.get(
                    "accion",
                    "editar"
                )
            ).strip().lower()

            if accion == "editar":
                editar_subseccion(
                    subseccion_id,
                    datos.get(
                        "nombre",
                        ""
                    ),
                    datos.get(
                        "descripcion",
                        ""
                    ),
                )

            elif accion == "desactivar":
                cambiar_estado_subseccion(
                    subseccion_id,
                    False,
                )

            elif accion == "reactivar":
                cambiar_estado_subseccion(
                    subseccion_id,
                    True,
                )

            else:
                raise ValueError(
                    "Acción de subsección no válida."
                )

        return jsonify(
            {
                "ok":
                    True
            }
        )

    except ValueError as error:
        return jsonify(
            {
                "ok":
                    False,
                "error":
                    str(
                        error
                    ),
            }
        ), 400

    except Exception as error:
        return jsonify(
            {
                "ok":
                    False,
                "error":
                    f"No fue posible actualizar la subsección: {error}",
            }
        ), 500


@repuestos_bp.route(
    "/api/inventario/materiales",
    methods=["POST"],
)
def api_crear_material_inventario():
    datos = request.get_json(
        silent=True
    ) or {}

    try:
        seccion_id = int(
            datos.get(
                "seccion_id"
            )
        )

        subseccion_valor = datos.get(
            "subseccion_id"
        )

        subseccion_id = (
            int(
                subseccion_valor
            )
            if subseccion_valor
            not in (
                None,
                "",
                0,
                "0",
            )
            else None
        )

        crear_material_inventario(
            codigo_sap=
                datos.get(
                    "codigo_sap",
                    ""
                ),

            descripcion=
                datos.get(
                    "descripcion",
                    ""
                ),

            unidad_medida=
                datos.get(
                    "unidad_medida",
                    ""
                ),

            seccion_id=
                seccion_id,

            subseccion_id=
                subseccion_id,

            stock_minimo=
                convertir_numero(
                    datos.get(
                        "stock_minimo"
                    )
                ),

            stock_objetivo=
                convertir_numero(
                    datos.get(
                        "stock_objetivo"
                    )
                ),

            observaciones=
                datos.get(
                    "observaciones",
                    ""
                ),

            origen_dato=
                datos.get(
                    "origen_dato",
                    "SAP"
                ),
        )

        return jsonify(
            {
                "ok":
                    True
            }
        )

    except (
        TypeError,
        ValueError,
    ) as error:
        return jsonify(
            {
                "ok":
                    False,
                "error":
                    str(
                        error
                    )
                    or "Revise la información del material.",
            }
        ), 400

    except Exception as error:
        return jsonify(
            {
                "ok":
                    False,
                "error":
                    f"No fue posible agregar el material: {error}",
            }
        ), 500


@repuestos_bp.route(
    "/api/inventario/materiales/<int:inventario_id>",
    methods=["PUT", "DELETE"],
)
def api_gestionar_material_inventario(
    inventario_id,
):
    datos = request.get_json(
        silent=True
    ) or {}

    try:
        if request.method == "DELETE":
            eliminar_material_inventario(
                inventario_id
            )

        else:
            accion = str(
                datos.get(
                    "accion",
                    "editar"
                )
            ).strip().lower()

            if accion == "editar":
                seccion_id = int(
                    datos.get(
                        "seccion_id"
                    )
                )

                subseccion_valor = datos.get(
                    "subseccion_id"
                )

                subseccion_id = (
                    int(
                        subseccion_valor
                    )
                    if subseccion_valor
                    not in (
                        None,
                        "",
                        0,
                        "0",
                    )
                    else None
                )

                editar_material_inventario(
                    inventario_id=
                        inventario_id,

                    seccion_id=
                        seccion_id,

                    subseccion_id=
                        subseccion_id,

                    stock_minimo=
                        convertir_numero(
                            datos.get(
                                "stock_minimo"
                            )
                        ),

                    stock_objetivo=
                        convertir_numero(
                            datos.get(
                                "stock_objetivo"
                            )
                        ),

                    observaciones=
                        datos.get(
                            "observaciones",
                            ""
                        ),
                )

            elif accion == "desactivar":
                cambiar_estado_material_inventario(
                    inventario_id,
                    False,
                )

            elif accion == "reactivar":
                cambiar_estado_material_inventario(
                    inventario_id,
                    True,
                )

            else:
                raise ValueError(
                    "Acción de material no válida."
                )

        return jsonify(
            {
                "ok":
                    True
            }
        )

    except (
        TypeError,
        ValueError,
    ) as error:
        return jsonify(
            {
                "ok":
                    False,
                "error":
                    str(
                        error
                    )
                    or "Revise la información del material.",
            }
        ), 400

    except Exception as error:
        return jsonify(
            {
                "ok":
                    False,
                "error":
                    f"No fue posible actualizar el material: {error}",
            }
        ), 500


@repuestos_bp.route(
    "/api/materiales"
)
def api_materiales():
    texto = request.args.get(
        "q",
        "",
    )

    return jsonify(
        buscar_materiales(
            texto
        )
    )


@repuestos_bp.route(
    "/api/material/<codigo_sap>/pendiente"
)
def api_material_pendiente(
    codigo_sap,
):
    pendientes = buscar_pedidos_pendientes(
        codigo_sap
    )

    return jsonify(
        {
            "pendiente":
                pendientes[0]
                if pendientes
                else None,

            "pendientes":
                pendientes,
        }
    )


@repuestos_bp.route(
    "/api/solicitudes",
    methods=["POST"],
)
def api_crear_solicitud():
    inicializar_bd()

    datos = request.get_json(
        silent=True
    ) or {}

    items = datos.get(
        "items",
        []
    )

    try:
        resultado = crear_solicitud(
            items
        )

        archivo = generar_excel_corporativo(
            resultado[
                "id"
            ]
        )

        resultado[
            "archivo"
        ] = archivo.name

        resultado[
            "archivo_url"
        ] = url_for(
            "repuestos.descargar_solicitud",
            solicitud_id=resultado[
                "id"
            ],
        )

        return jsonify(
            {
                "ok":
                    True,

                "solicitud":
                    resultado,
            }
        )

    except ValueError as error:
        return jsonify(
            {
                "ok":
                    False,

                "error":
                    str(
                        error
                    ),
            }
        ), 400

    except Exception as error:
        return jsonify(
            {
                "ok":
                    False,

                "error":
                    f"No fue posible guardar la solicitud: {error}",
            }
        ), 500



@repuestos_bp.route(
    "/solicitud/<int:solicitud_id>/descargar"
)
def descargar_solicitud(
    solicitud_id,
):
    conexion = conectar_bd()

    try:
        fila = conexion.execute(
            """
            SELECT
                archivo_generado,
                codigo_solicitud
            FROM solicitudes
            WHERE id = ?
            """,
            (
                solicitud_id,
            ),
        ).fetchone()

    finally:
        conexion.close()

    if (
        fila is None
        or not fila[
            "archivo_generado"
        ]
    ):
        return (
            "El archivo solicitado no existe.",
            404,
        )

    ruta = Path(
        fila[
            "archivo_generado"
        ]
    )

    if not ruta.exists():
        return (
            "El archivo solicitado no existe.",
            404,
        )

    return send_file(
        ruta,
        as_attachment=True,
        download_name=ruta.name,
    )
