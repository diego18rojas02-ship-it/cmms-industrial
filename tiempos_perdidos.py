from __future__ import annotations

import calendar
import os
import re
import shutil
import unicodedata
import uuid
from collections import defaultdict
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

from flask import (
    Blueprint,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from werkzeug.utils import secure_filename


tiempos_perdidos_bp = Blueprint(
    "tiempos_perdidos",
    __name__,
    url_prefix="/tiempos-perdidos",
)


def _openpyxl():
    """
    Importación diferida de openpyxl.

    Esto permite que el CMMS completo inicie aunque openpyxl todavía no
    esté instalado. La librería solo se exige al procesar un Excel.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.comments import Comment
        from openpyxl.utils.datetime import from_excel
    except ModuleNotFoundError as error:
        raise RuntimeError(
            "Falta instalar openpyxl. Ejecute: "
            "python -m pip install -r requirements.txt"
        ) from error

    return load_workbook, Comment, from_excel

EXTENSIONES_PERMITIDAS = {".xlsx", ".xlsm"}

PLANTILLA_TIEMPOS = "TIEMPOS TAPA PLASTICA.xlsx"
PLANTILLA_PARETO = "PARETO 80-20 TAPAS.xlsx"


def _excel_com():
    try:
        import pythoncom
        import win32com.client
    except ModuleNotFoundError as error:
        raise RuntimeError(
            "Falta instalar pywin32. Ejecute: python -m pip install pywin32"
        ) from error

    pythoncom.CoInitialize()

    excel = win32com.client.DispatchEx(
        "Excel.Application"
    )
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    excel.EnableEvents = False

    return excel, pythoncom


def _cerrar_excel_com(
    excel,
    pythoncom,
) -> None:
    try:
        excel.Quit()
    finally:
        pythoncom.CoUninitialize()


def _celda_com(
    hoja,
    fila: int,
    columna: int,
):
    return hoja.Cells(
        fila,
        columna,
    )


def _normalizar_nombre_hoja(
    nombre: str,
) -> str:
    return normalizar(
        nombre
    )


def _buscar_hoja_com(
    workbook,
    nombre: str,
):
    objetivo = _normalizar_nombre_hoja(
        nombre
    )

    for indice in range(
        1,
        workbook.Worksheets.Count + 1,
    ):
        hoja = workbook.Worksheets(
            indice
        )

        if (
            _normalizar_nombre_hoja(
                hoja.Name
            )
            == objetivo
        ):
            return hoja

    raise ValueError(
        f"No se encontró la hoja '{nombre}'."
    )


def _buscar_columnas_maquinas_com(
    hoja,
    fila_encabezado: int = 8,
) -> dict[str, int]:
    resultado = {}

    ultima_columna = hoja.Cells(
        fila_encabezado,
        hoja.Columns.Count,
    ).End(
        -4159
    ).Column

    for columna in range(
        1,
        ultima_columna + 1,
    ):
        valor = _celda_com(
            hoja,
            fila_encabezado,
            columna,
        ).Value

        maquina = resolver_maquina(
            valor
        )

        if maquina in ORDEN_MAQUINAS:
            resultado[
                maquina
            ] = columna

    faltantes = [
        maquina
        for maquina in ORDEN_MAQUINAS
        if maquina not in resultado
    ]

    if faltantes:
        raise ValueError(
            "El formato no contiene columnas para: "
            + ", ".join(
                faltantes
            )
        )

    return resultado


def _filas_dias_com(
    hoja,
) -> dict[int, int]:
    resultado = {}

    ultima_fila = hoja.Cells(
        hoja.Rows.Count,
        1,
    ).End(
        -4162
    ).Row

    for fila in range(
        9,
        ultima_fila + 1,
    ):
        valor = _celda_com(
            hoja,
            fila,
            1,
        ).Value

        dia = dia_desde_valor(
            valor
        )

        if (
            dia is not None
            and 1 <= dia <= 31
        ):
            resultado[
                dia
            ] = fila

    if not resultado:
        raise ValueError(
            f"No se encontraron días en la hoja '{hoja.Name}'."
        )

    return resultado


def _poner_comentario_com(
    celda,
    texto: str,
) -> None:
    try:
        if celda.Comment is not None:
            celda.Comment.Delete()
    except Exception:
        pass

    try:
        celda.AddComment(
            texto
        )
    except Exception:
        pass


def _llenar_matriz_com(
    hoja,
    datos: dict[
        tuple[int, str],
        dict[str, Any],
    ],
) -> None:
    columnas = _buscar_columnas_maquinas_com(
        hoja
    )

    filas_dias = _filas_dias_com(
        hoja
    )

    for fila in filas_dias.values():
        for columna in columnas.values():
            celda = _celda_com(
                hoja,
                fila,
                columna,
            )

            celda.Value = 0

            try:
                if celda.Comment is not None:
                    celda.Comment.Delete()
            except Exception:
                pass

    for (
        dia,
        maquina,
    ), grupo in datos.items():
        if dia not in filas_dias:
            continue

        if maquina not in columnas:
            continue

        celda = _celda_com(
            hoja,
            filas_dias[
                dia
            ],
            columnas[
                maquina
            ],
        )

        celda.Value = grupo[
            "minutos"
        ]

        _poner_comentario_com(
            celda,
            texto_comentario(
                grupo[
                    "descripciones"
                ]
            ),
        )


def _ultima_fila_total_com(
    hoja,
    columna: int,
    fila_inicio: int,
) -> int | None:
    ultima = hoja.Cells(
        hoja.Rows.Count,
        columna,
    ).End(
        -4162
    ).Row

    for fila in range(
        fila_inicio,
        ultima + 1,
    ):
        formula = _celda_com(
            hoja,
            fila,
            columna,
        ).Formula

        if (
            isinstance(
                formula,
                str,
            )
            and formula.upper().startswith(
                "=SUM("
            )
        ):
            return fila

    return None


def _asegurar_filas_com(
    hoja,
    fila_inicio: int,
    fila_total: int,
    cantidad: int,
) -> int:
    capacidad = max(
        0,
        fila_total - fila_inicio
    )

    extra = max(
        0,
        cantidad - capacidad
    )

    if extra <= 0:
        return fila_total

    for _ in range(
        extra
    ):
        hoja.Rows(
            fila_total
        ).Insert()

        hoja.Rows(
            fila_inicio
        ).Copy()

        hoja.Rows(
            fila_total
        ).PasteSpecial(
            -4122
        )

        fila_total += 1

    return fila_total


def _llenar_trabajos_realizados_com(
    workbook,
    fallas_por_maquina: dict[
        str,
        list[dict[str, Any]],
    ],
) -> None:
    for maquina in ORDEN_MAQUINAS:
        hoja = _buscar_hoja_com(
            workbook,
            _NOMBRE_HOJA_TRABAJO(
                maquina
            ),
        )

        fila_inicio = 8

        fila_total = _ultima_fila_total_com(
            hoja,
            4,
            fila_inicio,
        )

        if fila_total is None:
            fila_total = max(
                hoja.UsedRange.Rows.Count + 1,
                fila_inicio + 1,
            )

        fallas = fallas_por_maquina.get(
            maquina,
            [],
        )

        fila_total = _asegurar_filas_com(
            hoja,
            fila_inicio,
            fila_total,
            len(
                fallas
            ),
        )

        if fila_total > fila_inicio:
            hoja.Range(
                hoja.Cells(
                    fila_inicio,
                    1,
                ),
                hoja.Cells(
                    fila_total - 1,
                    5,
                ),
            ).ClearContents()

        for indice, falla in enumerate(
            fallas,
            start=fila_inicio,
        ):
            _celda_com(
                hoja,
                indice,
                1,
            ).Value = falla[
                "falla_no"
            ]

            _celda_com(
                hoja,
                indice,
                2,
            ).Value = falla[
                "descripcion"
            ]

            _celda_com(
                hoja,
                indice,
                3,
            ).Value = falla[
                "trabajo_realizado"
            ]

            _celda_com(
                hoja,
                indice,
                4,
            ).Value = falla[
                "minutos"
            ]

            _celda_com(
                hoja,
                indice,
                5,
            ).Value = falla[
                "eventos"
            ]

        ultima = (
            fila_inicio
            + len(
                fallas
            )
            - 1
        )

        if not fallas:
            ultima = fila_inicio

        _celda_com(
            hoja,
            fila_total,
            4,
        ).Formula = (
            f"=SUM(D{fila_inicio}:D{ultima})"
        )


def _NOMBRE_HOJA_TRABAJO(
    maquina: str,
) -> str:
    nombres = {
        "CCM1": "Trabajos Realizados CCM1",
        "CCM2": "Trabajos Realizados CCM2",
        "CCM3": "Trabajos Realizados CCM3",
        "CCM4": "Trabajos Realizados CCM64-1",
        "CCM5": "Trabajos Realizados CCM64-2",
        "CCM6": "Trabajos Realizados CCM 06",
        "CCM7": "Trabajos Realizados CCM 07",
        "CCM8": "Trabajos Realizados CCM 08",
        "PMV2": "Trabajos Realizados PMV 02",
        "PMV4": "Trabajos Realizados PMV 04",
    }

    return nombres[
        maquina
    ]


def _llenar_fallas_tapas_com(
    hoja,
    fallas: list[dict[str, Any]],
) -> None:
    fila_inicio = 4
    fila_total = 41

    fila_total = _asegurar_filas_com(
        hoja,
        fila_inicio,
        fila_total,
        len(
            fallas
        ),
    )

    hoja.Range(
        hoja.Cells(
            fila_inicio,
            2,
        ),
        hoja.Cells(
            fila_total - 1,
            7,
        ),
    ).ClearContents()

    posiciones = defaultdict(
        list
    )

    for fila, falla in enumerate(
        fallas,
        start=fila_inicio,
    ):
        _celda_com(
            hoja,
            fila,
            2,
        ).Value = falla[
            "falla_no"
        ]

        _celda_com(
            hoja,
            fila,
            3,
        ).Value = falla[
            "maquina"
        ]

        _celda_com(
            hoja,
            fila,
            4,
        ).Value = falla[
            "descripcion"
        ]

        _celda_com(
            hoja,
            fila,
            5,
        ).Value = falla[
            "minutos"
        ]

        _celda_com(
            hoja,
            fila,
            6,
        ).Value = falla[
            "eventos"
        ]

        posiciones[
            falla[
                "maquina"
            ]
        ].append(
            fila
        )

    for filas in posiciones.values():
        primera = filas[0]
        ultima = filas[-1]

        _celda_com(
            hoja,
            primera,
            7,
        ).Formula = (
            f"=SUM(E{primera}:E{ultima})"
        )

    ultima_datos = (
        fila_inicio
        + len(
            fallas
        )
        - 1
    )

    if not fallas:
        ultima_datos = fila_inicio

    _celda_com(
        hoja,
        fila_total,
        4,
    ).Value = (
        "TOTAL TIEMPOS PERDIDOS AREA"
    )

    _celda_com(
        hoja,
        fila_total,
        5,
    ).Formula = (
        f"=SUM(E{fila_inicio}:E{ultima_datos})"
    )


def _llenar_datos_pareto_com(
    hoja,
    fallas: list[dict[str, Any]],
) -> None:
    fila_inicio = 4
    fila_total = 40

    ordenadas = sorted(
        fallas,
        key=lambda item:
            item[
                "minutos"
            ],
        reverse=True,
    )

    fila_total = _asegurar_filas_com(
        hoja,
        fila_inicio,
        fila_total,
        len(
            ordenadas
        ),
    )

    hoja.Range(
        hoja.Cells(
            fila_inicio,
            4,
        ),
        hoja.Cells(
            fila_total - 1,
            8,
        ),
    ).ClearContents()

    for numero, falla in enumerate(
        ordenadas,
        start=1,
    ):
        fila = (
            fila_inicio
            + numero
            - 1
        )

        _celda_com(
            hoja,
            fila,
            4,
        ).Value = numero

        _celda_com(
            hoja,
            fila,
            5,
        ).Value = falla[
            "maquina"
        ]

        _celda_com(
            hoja,
            fila,
            6,
        ).Value = falla[
            "descripcion"
        ]

        _celda_com(
            hoja,
            fila,
            7,
        ).Value = falla[
            "minutos"
        ]

        _celda_com(
            hoja,
            fila,
            8,
        ).Value = falla[
            "eventos"
        ]

    ultima = (
        fila_inicio
        + len(
            ordenadas
        )
        - 1
    )

    if not ordenadas:
        ultima = fila_inicio

    _celda_com(
        hoja,
        fila_total,
        6,
    ).Value = (
        "TOTAL TIEMPOS PERDIDOS AREA"
    )

    _celda_com(
        hoja,
        fila_total,
        7,
    ).Formula = (
        f"=SUM(G{fila_inicio}:G{ultima})"
    )

    hoja.Range(
        "J4"
    ).Value = "PARETO 80/20"

    hoja.Range(
        "J5"
    ).Value = "EVENTOS 20%"

    hoja.Range(
        "K5"
    ).Formula = (
        f"=COUNTA(D{fila_inicio}:D{ultima})*20%"
    )

    hoja.Range(
        "J6"
    ).Value = "TIEMPO 80 %"

    hoja.Range(
        "K6"
    ).Formula = (
        f"=G{fila_total}*80%"
    )


CLASES_ORDEN = {
    "PM01A": "correctivo",
    "PM01B": "preventivo",
    "PM02": "preventivo",
    "PM02A": "preventivo",
    "PM02B": "preventivo",
    "PMV02": "preventivo",
    "PM03": "cambio",
}

HOJAS_DESTINO_TIEMPOS = {
    "correctivo": "tiempos Perdidos TP IBERPLAST",
    "preventivo": "tiempos Mtto Prev TP",
    "cambio": "tiempos Cambios TP",
}


ALIAS_MAQUINAS = {
    "CCM1": [
        "TERMOCOMPRESORA SACMI Nº1",
        "TERMOCOMPRESORA SACMI Nº 1",
        "TERMOCOMPRESORA SACMI N°1",
        "TERMOCOMPRESORA SACMI N° 1",
        "TERMOCOMPRESORA SACMI NO 1",
        "TERMOCOMPRESORA SACMI 1",
        "CCM1",
        "CCM 1",
        "CCM01",
        "CCM 01",
    ],

    "CCM2": [
        "TERMOCOMPRESORA SACMI Nº2",
        "TERMOCOMPRESORA SACMI Nº 2",
        "TERMOCOMPRESORA SACMI N°2",
        "TERMOCOMPRESORA SACMI N° 2",
        "TERMOCOMPRESORA SACMI NO 2",
        "TERMOCOMPRESORA SACMI 2",
        "CCM2",
        "CCM 2",
        "CCM02",
        "CCM 02",
    ],

    "CCM3": [
        "TERMOCOMPRESORA SACMI Nº3",
        "TERMOCOMPRESORA SACMI Nº 3",
        "TERMOCOMPRESORA SACMI N°3",
        "TERMOCOMPRESORA SACMI N° 3",
        "TERMOCOMPRESORA SACMI NO 3",
        "TERMOCOMPRESORA SACMI 3",
        "CCM3",
        "CCM 3",
        "CCM03",
        "CCM 03",
    ],

    "CCM4": [
        "TERMOCOMPRESORA SACMI Nº4",
        "TERMOCOMPRESORA SACMI Nº 4",
        "TERMOCOMPRESORA SACMI N°4",
        "TERMOCOMPRESORA SACMI N° 4",
        "TERMOCOMPRESORA SACMI NO 4",
        "TERMOCOMPRESORA SACMI 4",
        "CCM4",
        "CCM 4",
        "CCM04",
        "CCM 04",
        "CCM64-1",
        "CCM 64-1",
        "CCM64 - 1",
        "CCM 64 - 1",
    ],

    "CCM5": [
        "TERMOCOMPRESORA SACMI Nº5",
        "TERMOCOMPRESORA SACMI Nº 5",
        "TERMOCOMPRESORA SACMI N°5",
        "TERMOCOMPRESORA SACMI N° 5",
        "TERMOCOMPRESORA SACMI NO 5",
        "TERMOCOMPRESORA SACMI 5",
        "CCM5",
        "CCM 5",
        "CCM05",
        "CCM 05",
        "CCM64-2",
        "CCM 64-2",
        "CCM64 - 2",
        "CCM 64 - 2",
    ],

    "CCM6": [
        "TERMOCOMPRESORA SACMI Nº6",
        "TERMOCOMPRESORA SACMI Nº 6",
        "TERMOCOMPRESORA SACMI N°6",
        "TERMOCOMPRESORA SACMI N° 6",
        "TERMOCOMPRESORA SACMI NO 6",
        "TERMOCOMPRESORA SACMI 6",
        "CCM6",
        "CCM 6",
        "CCM06",
        "CCM 06",
    ],

    "CCM7": [
        "TERMOCOMPRESORA SACMI Nº7",
        "TERMOCOMPRESORA SACMI Nº 7",
        "TERMOCOMPRESORA SACMI N°7",
        "TERMOCOMPRESORA SACMI N° 7",
        "TERMOCOMPRESORA SACMI NO 7",
        "TERMOCOMPRESORA SACMI 7",
        "CCM7",
        "CCM 7",
        "CCM07",
        "CCM 07",
    ],

    "CCM8": [
        "TERMOCOMPRESORA SACMI Nº8",
        "TERMOCOMPRESORA SACMI Nº 8",
        "TERMOCOMPRESORA SACMI N°8",
        "TERMOCOMPRESORA SACMI N° 8",
        "TERMOCOMPRESORA SACMI NO 8",
        "TERMOCOMPRESORA SACMI 8",
        "CCM8",
        "CCM 8",
        "CCM08",
        "CCM 08",
    ],

    "PMV2": [
        "SACMI PMV N° 2",
        "SACMI PMV Nº 2",
        "SACMI PMV N°2",
        "SACMI PMV Nº2",
        "SACMI PMV NO 2",
        "SACMI PMV 2",
        "PMV2",
        "PMV 2",
        "PMV02",
        "PMV 02",
    ],

    "PMV4": [
        "SACMI PMV N° 4",
        "SACMI PMV Nº 4",
        "SACMI PMV N°4",
        "SACMI PMV Nº4",
        "SACMI PMV NO 4",
        "SACMI PMV 4",
        "PMV4",
        "PMV 4",
        "PMV04",
        "PMV 04",
    ],
}


ORDEN_MAQUINAS = [
    "CCM1",
    "CCM2",
    "CCM3",
    "CCM4",
    "CCM5",
    "CCM6",
    "CCM7",
    "CCM8",
    "PMV2",
    "PMV4",
]

ENCABEZADOS = {
    "dia": {
        "DIA",
        "FECHA",
    },
    "equipo": {
        "EQUIPO",
        "MAQUINA",
    },
    "clase": {
        "CLASE DE ORDEN",
        "CLASE ORDEN",
    },
    "descripcion": {
        "DESCRIPCION PARADA",
        "DESCRIPCION DE PARADA",
        "DESCRIPCION",
    },
    "trabajo": {
        "TRABAJO REALIZADO",
    },
    "estado": {
        "ESTADO",
        "ESTADO TRABAJO",
        "ESTADO DE TRABAJO",
        "ESTADO DEL TRABAJO",
    },
    "tiempo": {
        "TIEMPO PERDIDO MIN",
        "TIEMPO PERDIDO",
        "TIEMPO PERDIDO MINUTOS",
    },
}


def ruta_base() -> Path:
    return Path(
        current_app.root_path
    )


def carpeta_uploads() -> Path:
    ruta = ruta_base() / "uploads_tiempos"
    ruta.mkdir(
        parents=True,
        exist_ok=True,
    )
    return ruta


def carpeta_salidas() -> Path:
    ruta = ruta_base() / "salidas"
    ruta.mkdir(
        parents=True,
        exist_ok=True,
    )
    return ruta


def carpeta_formatos() -> Path:
    ruta = ruta_base() / "formatos"
    ruta.mkdir(
        parents=True,
        exist_ok=True,
    )
    return ruta


def normalizar(texto: Any) -> str:
    if texto is None:
        return ""

    valor = str(texto).strip().upper()

    valor = unicodedata.normalize(
        "NFKD",
        valor,
    )

    valor = "".join(
        caracter
        for caracter in valor
        if not unicodedata.combining(
            caracter
        )
    )

    valor = valor.replace(
        "º",
        " ",
    ).replace(
        "°",
        " ",
    ).replace(
        "N°",
        "N ",
    )

    valor = re.sub(
        r"[\r\n\t]+",
        " ",
        valor,
    )

    valor = re.sub(
        r"[^A-Z0-9]+",
        " ",
        valor,
    )

    valor = re.sub(
        r"\s+",
        " ",
        valor,
    ).strip()

    return valor


def construir_mapa_alias() -> dict[str, str]:
    resultado = {}

    for codigo, alias in ALIAS_MAQUINAS.items():
        resultado[
            normalizar(
                codigo
            )
        ] = codigo

        for nombre in alias:
            resultado[
                normalizar(
                    nombre
                )
            ] = codigo

    return resultado


MAPA_ALIAS = construir_mapa_alias()


def resolver_maquina(
    nombre: Any,
) -> str | None:
    limpio = normalizar(
        nombre
    )

    if not limpio:
        return None

    if limpio in MAPA_ALIAS:
        return MAPA_ALIAS[
            limpio
        ]

    coincidencia = re.search(
        r"\bTERMOCOMPRESORA\s+SACMI\s+(?:N|NO)?\s*0?([1-8])\b",
        limpio,
    )

    if coincidencia:
        return (
            "CCM"
            + str(
                int(
                    coincidencia.group(
                        1
                    )
                )
            )
        )

    if re.search(
        r"\bCCM\s*64\s*1\b",
        limpio,
    ):
        return "CCM4"

    if re.search(
        r"\bCCM\s*64\s*2\b",
        limpio,
    ):
        return "CCM5"

    coincidencia = re.search(
        r"\bCCM\s*0?([1-8])\b",
        limpio,
    )

    if coincidencia:
        return (
            "CCM"
            + str(
                int(
                    coincidencia.group(
                        1
                    )
                )
            )
        )

    coincidencia = re.search(
        r"\b(?:SACMI\s+)?PMV\s+(?:N|NO)?\s*0?(2|4)\b",
        limpio,
    )

    if coincidencia:
        return (
            "PMV"
            + str(
                int(
                    coincidencia.group(
                        1
                    )
                )
            )
        )

    return None


def valor_minutos(
    valor: Any,
) -> int:
    if valor in (
        None,
        "",
    ):
        return 0

    if isinstance(
        valor,
        bool,
    ):
        return 0

    if isinstance(
        valor,
        (
            int,
            float,
        ),
    ):
        return int(
            round(
                float(valor)
            )
        )

    texto = str(
        valor
    ).strip()

    texto = texto.replace(
        ".",
        "",
    ).replace(
        ",",
        ".",
    )

    try:
        return int(
            round(
                float(texto)
            )
        )
    except ValueError:
        return 0


def dia_desde_valor(
    valor: Any,
) -> int | None:
    _, _, from_excel = _openpyxl()
    if valor is None:
        return None

    if isinstance(
        valor,
        datetime,
    ):
        return valor.day

    if isinstance(
        valor,
        date,
    ):
        return valor.day

    if isinstance(
        valor,
        (
            int,
            float,
        ),
    ):
        numero = int(
            valor
        )

        if 1 <= numero <= 31:
            return numero

        try:
            fecha = from_excel(
                valor
            )

            return fecha.day
        except Exception:
            return None

    texto = str(
        valor
    ).strip()

    if not texto:
        return None

    if texto.isdigit():
        numero = int(
            texto
        )

        if 1 <= numero <= 31:
            return numero

    for formato in (
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%d-%m-%Y",
    ):
        try:
            return datetime.strptime(
                texto,
                formato,
            ).day
        except ValueError:
            continue

    return None


def mes_anio_desde_fila(
    fila: tuple[Any, ...],
    mapa: dict[str, int],
) -> tuple[int | None, int | None]:
    mes = None
    anio = None

    for indice, valor in enumerate(
        fila,
        start=1,
    ):
        pass

    return mes, anio


def buscar_hoja_plastica(
    workbook,
):
    candidatas = []

    for hoja in workbook.worksheets:
        nombre = normalizar(
            hoja.title
        )

        if (
            "PLAST" in nombre
            and "METAL" not in nombre
        ):
            candidatas.append(
                hoja
            )

    if not candidatas:
        raise ValueError(
            "No se encontró una hoja de Plástica en el archivo de Entrega de Turno."
        )

    candidatas.sort(
        key=lambda hoja:
            hoja.max_row,
        reverse=True,
    )

    return candidatas[0]


def buscar_fila_encabezados(
    hoja,
) -> tuple[int, dict[str, int], dict[str, int]]:
    for numero_fila in range(
        1,
        min(
            hoja.max_row,
            80,
        )
        + 1,
    ):
        mapa = {}
        mapa_extra = {}

        for celda in hoja[
            numero_fila
        ]:
            encabezado = normalizar(
                celda.value
            )

            if not encabezado:
                continue

            for campo, opciones in ENCABEZADOS.items():
                if encabezado in opciones:
                    mapa[
                        campo
                    ] = celda.column

            if (
                encabezado.startswith(
                    "ESTADO DE TRABAJO"
                )
                or encabezado.startswith(
                    "ESTADO TRABAJO"
                )
                or encabezado.startswith(
                    "ESTADO DEL TRABAJO"
                )
            ):
                mapa[
                    "estado"
                ] = celda.column

            if encabezado == "MES":
                mapa_extra[
                    "mes"
                ] = celda.column

            if encabezado in {
                "ANO",
                "ANIO",
            }:
                mapa_extra[
                    "anio"
                ] = celda.column

        obligatorios = {
            "dia",
            "equipo",
            "clase",
            "descripcion",
            "trabajo",
            "tiempo",
        }

        if obligatorios.issubset(
            mapa.keys()
        ):
            return (
                numero_fila,
                mapa,
                mapa_extra,
            )

    raise ValueError(
        "No se encontró la fila de encabezados. "
        "Se requieren: Día, Equipo, Clase de orden, "
        "Descripción parada, Trabajo realizado y Tiempo perdido."
    )


def nombre_mes_a_numero(
    valor: Any,
) -> int | None:
    if valor in (
        None,
        "",
    ):
        return None

    if isinstance(
        valor,
        (
            int,
            float,
        ),
    ):
        numero = int(
            valor
        )

        if 1 <= numero <= 12:
            return numero

    limpio = normalizar(
        valor
    )

    meses = {
        "ENERO": 1,
        "FEBRERO": 2,
        "MARZO": 3,
        "ABRIL": 4,
        "MAYO": 5,
        "JUNIO": 6,
        "JULIO": 7,
        "AGOSTO": 8,
        "SEPTIEMBRE": 9,
        "SETIEMBRE": 9,
        "OCTUBRE": 10,
        "NOVIEMBRE": 11,
        "DICIEMBRE": 12,
    }

    return meses.get(
        limpio
    )


def leer_entrega_turno(
    ruta: Path,
) -> dict[str, Any]:
    load_workbook, _, _ = _openpyxl()

    workbook = load_workbook(
        ruta,
        data_only=True,
        read_only=True,
    )

    try:
        hoja = buscar_hoja_plastica(
            workbook
        )

        fila_encabezados, columnas, extras = (
            buscar_fila_encabezados(
                hoja
            )
        )

        registros = []
        clases_ignoradas = set()
        mes_detectado = None
        anio_detectado = None
        filas_vacias_consecutivas = 0

        for numero_fila in range(
            fila_encabezados + 1,
            hoja.max_row + 1,
        ):
            valor_equipo = hoja.cell(
                numero_fila,
                columnas[
                    "equipo"
                ],
            ).value

            if (
                valor_equipo is None
                or str(
                    valor_equipo
                ).strip() == ""
            ):
                filas_vacias_consecutivas += 1

                if filas_vacias_consecutivas >= 10:
                    break

                continue

            filas_vacias_consecutivas = 0

            maquina = resolver_maquina(
                valor_equipo
            )

            if (
                maquina is None
                or maquina not in ORDEN_MAQUINAS
            ):
                continue

            clase_original = hoja.cell(
                numero_fila,
                columnas[
                    "clase"
                ],
            ).value

            clase = normalizar(
                clase_original
            ).replace(
                " ",
                "",
            )

            if clase not in CLASES_ORDEN:
                if clase:
                    clases_ignoradas.add(
                        str(
                            clase_original
                        )
                    )

                continue

            dia = dia_desde_valor(
                hoja.cell(
                    numero_fila,
                    columnas[
                        "dia"
                    ],
                ).value
            )

            if dia is None:
                continue

            minutos = valor_minutos(
                hoja.cell(
                    numero_fila,
                    columnas[
                        "tiempo"
                    ],
                ).value
            )

            if minutos <= 0:
                continue

            descripcion = str(
                hoja.cell(
                    numero_fila,
                    columnas[
                        "descripcion"
                    ],
                ).value
                or ""
            ).strip()

            trabajo = str(
                hoja.cell(
                    numero_fila,
                    columnas[
                        "trabajo"
                    ],
                ).value
                or ""
            ).strip()

            estado = ""

            if "estado" in columnas:
                estado = str(
                    hoja.cell(
                        numero_fila,
                        columnas[
                            "estado"
                        ],
                    ).value
                    or ""
                ).strip()

            if (
                mes_detectado is None
                and "mes" in extras
            ):
                mes_detectado = nombre_mes_a_numero(
                    hoja.cell(
                        numero_fila,
                        extras[
                            "mes"
                        ],
                    ).value
                )

            if (
                anio_detectado is None
                and "anio" in extras
            ):
                valor_anio = hoja.cell(
                    numero_fila,
                    extras[
                        "anio"
                    ],
                ).value

                try:
                    anio_detectado = int(
                        valor_anio
                    )
                except (
                    TypeError,
                    ValueError,
                ):
                    pass

            registros.append(
                {
                    "fila_origen":
                        numero_fila,
                    "dia":
                        dia,
                    "equipo_origen":
                        str(
                            valor_equipo
                        ).strip(),
                    "maquina":
                        maquina,
                    "clase_orden":
                        clase,
                    "categoria":
                        CLASES_ORDEN[
                            clase
                        ],
                    "descripcion":
                        descripcion,
                    "trabajo_realizado":
                        trabajo,
                    "estado":
                        estado,
                    "minutos":
                        minutos,
                }
            )

        if not registros:
            raise ValueError(
                "No se encontraron registros válidos para las máquinas "
                "y clases de orden configuradas."
            )

        return {
            "hoja":
                hoja.title,
            "registros":
                registros,
            "maquinas_no_reconocidas":
                [],
            "clases_ignoradas":
                sorted(
                    clases_ignoradas
                ),
            "mes":
                mes_detectado,
            "anio":
                anio_detectado,
        }

    finally:
        workbook.close()


def nombre_formato_maquina(
    codigo: str,
) -> str:
    nombres = {
        "CCM1": "CCM 01",
        "CCM2": "CCM 02",
        "CCM3": "CCM 03",
        "CCM4": "CCM 04",
        "CCM5": "CCM 05",
        "CCM6": "CCM 06",
        "CCM7": "CCM 07",
        "CCM8": "CCM 08",
        "PMV2": "PMV 02",
        "PMV4": "PMV 04",
    }

    return nombres[
        codigo
    ]


def buscar_columnas_maquinas(
    hoja,
    fila_encabezado: int = 8,
) -> dict[str, int]:
    resultado = {}

    for celda in hoja[
        fila_encabezado
    ]:
        valor = normalizar(
            celda.value
        )

        if not valor:
            continue

        maquina = resolver_maquina(
            valor
        )

        if maquina in ORDEN_MAQUINAS:
            resultado[
                maquina
            ] = celda.column

    faltantes = [
        maquina
        for maquina in ORDEN_MAQUINAS
        if maquina not in resultado
    ]

    if faltantes:
        raise ValueError(
            "El formato no contiene columnas para: "
            + ", ".join(
                faltantes
            )
        )

    return resultado


def filas_dias_del_formato(
    hoja,
) -> dict[int, int]:
    resultado = {}

    for numero_fila in range(
        9,
        hoja.max_row + 1,
    ):
        valor = hoja.cell(
            numero_fila,
            1,
        ).value

        dia = dia_desde_valor(
            valor
        )

        if dia is None:
            continue

        if 1 <= dia <= 31:
            resultado[
                dia
            ] = numero_fila

    if not resultado:
        raise ValueError(
            f"No se encontraron días en la columna A de la hoja '{hoja.title}'."
        )

    return resultado


def _agregar_unico(
    lista: list[str],
    valor: str,
) -> None:
    limpio = str(
        valor
        or ""
    ).strip()

    if not limpio:
        return

    clave = normalizar(
        limpio
    )

    existentes = {
        normalizar(
            item
        )
        for item in lista
    }

    if clave not in existentes:
        lista.append(
            limpio
        )


def construir_agregados(
    registros: list[dict[str, Any]],
) -> dict[str, dict[tuple[int, str], dict[str, Any]]]:
    agregados = {
        "correctivo": {},
        "preventivo": {},
        "cambio": {},
    }

    for registro in registros:
        categoria = registro[
            "categoria"
        ]

        clave = (
            registro[
                "dia"
            ],
            registro[
                "maquina"
            ],
        )

        grupo = agregados[
            categoria
        ].setdefault(
            clave,
            {
                "minutos": 0,
                "descripciones": [],
            },
        )

        grupo[
            "minutos"
        ] += registro[
            "minutos"
        ]

        _agregar_unico(
            grupo[
                "descripciones"
            ],
            registro[
                "descripcion"
            ],
        )

    return agregados


def texto_comentario(
    descripciones: list[str],
) -> str:
    limpias = []

    for descripcion in descripciones:
        _agregar_unico(
            limpias,
            descripcion,
        )

    if not limpias:
        return "Sin descripción de parada."

    if len(
        limpias
    ) == 1:
        return limpias[0]

    return "\n".join(
        f"{indice}. {descripcion}"
        for indice, descripcion in enumerate(
            limpias,
            start=1,
        )
    )


def construir_fallas_correctivas(
    registros: list[dict[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    grupos: dict[
        tuple[str, str],
        dict[str, Any],
    ] = {}

    for registro in registros:
        if registro[
            "categoria"
        ] != "correctivo":
            continue

        descripcion = str(
            registro[
                "descripcion"
            ]
            or ""
        ).strip()

        clave_descripcion = normalizar(
            descripcion
        )

        if not clave_descripcion:
            clave_descripcion = (
                "SIN DESCRIPCION"
            )

        clave = (
            registro[
                "maquina"
            ],
            clave_descripcion,
        )

        grupo = grupos.setdefault(
            clave,
            {
                "maquina":
                    registro[
                        "maquina"
                    ],
                "descripcion":
                    descripcion
                    or "Sin descripción de parada",
                "minutos":
                    0,
                "trabajos_terminados":
                    [],
                "terminados":
                    0,
                "registros":
                    0,
            },
        )

        grupo[
            "minutos"
        ] += registro[
            "minutos"
        ]

        grupo[
            "registros"
        ] += 1

        estado = normalizar(
            registro.get(
                "estado",
                "",
            )
        )

        if "TERMINADO" in estado:
            grupo[
                "terminados"
            ] += 1

            _agregar_unico(
                grupo[
                    "trabajos_terminados"
                ],
                registro.get(
                    "trabajo_realizado",
                    "",
                ),
            )

    por_maquina = {
        maquina: []
        for maquina in ORDEN_MAQUINAS
    }

    for grupo in grupos.values():
        grupo[
            "eventos"
        ] = max(
            1,
            grupo[
                "terminados"
            ],
        )

        grupo[
            "trabajo_realizado"
        ] = "\n".join(
            grupo[
                "trabajos_terminados"
            ]
        )

        por_maquina[
            grupo[
                "maquina"
            ]
        ].append(
            grupo
        )

    for maquina in por_maquina:
        por_maquina[
            maquina
        ].sort(
            key=lambda item: (
                -item[
                    "minutos"
                ],
                normalizar(
                    item[
                        "descripcion"
                    ]
                ),
            )
        )

        for numero, falla in enumerate(
            por_maquina[
                maquina
            ],
            start=1,
        ):
            falla[
                "falla_no"
            ] = numero

    return por_maquina


def limpiar_matriz_tiempos(
    hoja,
    filas_dias: dict[int, int],
    columnas: dict[str, int],
) -> None:
    for fila in filas_dias.values():
        for columna in columnas.values():
            celda = hoja.cell(
                fila,
                columna,
            )

            celda.value = 0
            celda.comment = None


def llenar_matriz(
    hoja,
    datos: dict[tuple[int, str], dict[str, Any]],
) -> None:
    _, Comment, _ = _openpyxl()

    columnas = buscar_columnas_maquinas(
        hoja
    )

    filas_dias = filas_dias_del_formato(
        hoja
    )

    limpiar_matriz_tiempos(
        hoja,
        filas_dias,
        columnas,
    )

    for (
        dia,
        maquina,
    ), grupo in datos.items():
        if dia not in filas_dias:
            raise ValueError(
                f"El día {dia} no existe en la hoja '{hoja.title}'."
            )

        fila = filas_dias[
            dia
        ]

        columna = columnas[
            maquina
        ]

        celda = hoja.cell(
            fila,
            columna,
        )

        celda.value = grupo[
            "minutos"
        ]

        celda.comment = Comment(
            texto_comentario(
                grupo[
                    "descripciones"
                ]
            ),
            "CMMS Industrial",
        )


def buscar_hoja_por_nombre(
    workbook,
    nombre_esperado: str,
):
    if nombre_esperado in workbook.sheetnames:
        return workbook[
            nombre_esperado
        ]

    esperado = normalizar(
        nombre_esperado
    )

    for hoja in workbook.worksheets:
        if normalizar(
            hoja.title
        ) == esperado:
            return hoja

    raise ValueError(
        f"No se encontró la hoja '{nombre_esperado}'."
    )


def _hoja_trabajos_por_maquina(
    workbook,
    maquina: str,
):
    nombres = {
        "CCM1":
            "Trabajos Realizados CCM1",
        "CCM2":
            "Trabajos Realizados CCM2",
        "CCM3":
            "Trabajos Realizados CCM3",
        "CCM4":
            "Trabajos Realizados CCM64-1",
        "CCM5":
            "Trabajos Realizados CCM64-2",
        "CCM6":
            "Trabajos Realizados CCM 06",
        "CCM7":
            "Trabajos Realizados CCM 07",
        "CCM8":
            "Trabajos Realizados CCM 08",
        "PMV2":
            "Trabajos Realizados PMV 02",
        "PMV4":
            "Trabajos Realizados PMV 04",
    }

    nombre = nombres[
        maquina
    ]

    return buscar_hoja_por_nombre(
        workbook,
        nombre,
    )


def buscar_fila_total(
    hoja,
    columna: int,
    fila_inicio: int,
) -> int | None:
    for fila in range(
        fila_inicio,
        hoja.max_row + 1,
    ):
        valor = hoja.cell(
            fila,
            columna,
        ).value

        if (
            isinstance(
                valor,
                str,
            )
            and valor.startswith(
                "=SUM("
            )
        ):
            return fila

    return None


def copiar_altura_fila(
    hoja,
    fila_origen: int,
    fila_destino: int,
) -> None:
    altura = hoja.row_dimensions[
        fila_origen
    ].height

    if altura is not None:
        hoja.row_dimensions[
            fila_destino
        ].height = altura


def asegurar_filas_antes_de_total(
    hoja,
    fila_inicio: int,
    fila_total: int,
    cantidad_necesaria: int,
    fila_estilo: int,
    columnas: range,
) -> int:
    capacidad = max(
        0,
        fila_total - fila_inicio,
    )

    extra = max(
        0,
        cantidad_necesaria - capacidad,
    )

    if extra > 0:
        hoja.insert_rows(
            fila_total,
            amount=extra,
        )

        for fila in range(
            fila_total,
            fila_total + extra,
        ):
            copiar_estilo_fila(
                hoja,
                fila_estilo,
                fila,
                columnas,
            )

            copiar_altura_fila(
                hoja,
                fila_estilo,
                fila,
            )

        fila_total += extra

    return fila_total


def limpiar_datos_sin_tocar_combinadas(
    hoja,
    fila_inicio: int,
    fila_fin: int,
    columna_inicio: int,
    columna_fin: int,
) -> None:
    for fila in range(
        fila_inicio,
        fila_fin + 1,
    ):
        for columna in range(
            columna_inicio,
            columna_fin + 1,
        ):
            celda = hoja.cell(
                fila,
                columna,
            )

            if (
                celda.__class__.__name__
                == "MergedCell"
            ):
                continue

            celda.value = None


def llenar_trabajos_realizados(
    workbook,
    fallas_por_maquina: dict[
        str,
        list[dict[str, Any]],
    ],
) -> None:
    for maquina in ORDEN_MAQUINAS:
        hoja = _hoja_trabajos_por_maquina(
            workbook,
            maquina,
        )

        fila_inicio = 8

        fila_total = buscar_fila_total(
            hoja,
            4,
            fila_inicio,
        )

        if fila_total is None:
            fila_total = max(
                hoja.max_row + 1,
                fila_inicio + 1,
            )

        fallas = fallas_por_maquina.get(
            maquina,
            [],
        )

        fila_total = asegurar_filas_antes_de_total(
            hoja,
            fila_inicio,
            fila_total,
            len(
                fallas
            ),
            fila_inicio,
            range(
                1,
                6,
            ),
        )

        limpiar_datos_sin_tocar_combinadas(
            hoja,
            fila_inicio,
            fila_total - 1,
            1,
            5,
        )

        for indice, falla in enumerate(
            fallas,
            start=fila_inicio,
        ):
            copiar_estilo_fila(
                hoja,
                fila_inicio,
                indice,
                range(
                    1,
                    6,
                ),
            )

            copiar_altura_fila(
                hoja,
                fila_inicio,
                indice,
            )

            hoja.cell(
                indice,
                1,
            ).value = falla[
                "falla_no"
            ]

            hoja.cell(
                indice,
                2,
            ).value = falla[
                "descripcion"
            ]

            hoja.cell(
                indice,
                3,
            ).value = falla[
                "trabajo_realizado"
            ]

            hoja.cell(
                indice,
                4,
            ).value = falla[
                "minutos"
            ]

            hoja.cell(
                indice,
                5,
            ).value = falla[
                "eventos"
            ]

        ultima_fila = (
            fila_inicio
            + len(
                fallas
            )
            - 1
        )

        if not fallas:
            ultima_fila = fila_inicio

        hoja.cell(
            fila_total,
            4,
        ).value = (
            f"=SUM(D{fila_inicio}:D{ultima_fila})"
        )


def diligenciar_formato_tiempos(
    plantilla: Path,
    salida: Path,
    agregados: dict[
        str,
        dict[
            tuple[int, str],
            dict[str, Any],
        ],
    ],
    fallas_por_maquina: dict[
        str,
        list[dict[str, Any]],
    ],
) -> None:
    shutil.copy2(
        plantilla,
        salida,
    )

    excel, pythoncom = _excel_com()
    workbook = None

    try:
        workbook = excel.Workbooks.Open(
            str(
                salida.resolve()
            )
        )

        for categoria, nombre_hoja in HOJAS_DESTINO_TIEMPOS.items():
            hoja = _buscar_hoja_com(
                workbook,
                nombre_hoja,
            )

            _llenar_matriz_com(
                hoja,
                agregados[
                    categoria
                ],
            )

        _llenar_trabajos_realizados_com(
            workbook,
            fallas_por_maquina,
        )

        workbook.Save()

    finally:
        if workbook is not None:
            workbook.Close(
                SaveChanges=True
            )

        _cerrar_excel_com(
            excel,
            pythoncom,
        )


def limpiar_rango_valores(
    hoja,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> None:
    for fila in range(
        min_row,
        max_row + 1,
    ):
        for columna in range(
            min_col,
            max_col + 1,
        ):
            celda = hoja.cell(
                fila,
                columna,
            )

            if (
                celda.__class__.__name__
                == "MergedCell"
            ):
                continue

            celda.value = None


def fallas_para_pareto(
    fallas_por_maquina: dict[
        str,
        list[dict[str, Any]],
    ],
) -> list[dict[str, Any]]:
    resultado = []

    for maquina in (
        "CCM1",
        "CCM2",
        "CCM3",
        "CCM4",
        "CCM5",
        "CCM6",
        "CCM7",
        "CCM8",
    ):
        for falla in fallas_por_maquina.get(
            maquina,
            [],
        ):
            resultado.append(
                dict(
                    falla
                )
            )

    return resultado


def copiar_estilo_fila(
    hoja,
    fila_origen: int,
    fila_destino: int,
    columnas: range,
) -> None:
    for columna in columnas:
        origen = hoja.cell(
            fila_origen,
            columna,
        )

        destino = hoja.cell(
            fila_destino,
            columna,
        )

        if origen.has_style:
            destino._style = copy(
                origen._style
            )

        destino.font = copy(
            origen.font
        )

        destino.fill = copy(
            origen.fill
        )

        destino.border = copy(
            origen.border
        )

        destino.alignment = copy(
            origen.alignment
        )

        destino.number_format = (
            origen.number_format
        )

        destino.protection = copy(
            origen.protection
        )


def llenar_fallas_tapas(
    hoja,
    fallas: list[dict[str, Any]],
) -> None:
    fila_inicio = 4
    fila_total_original = 41

    for rango in list(
        hoja.merged_cells.ranges
    ):
        if (
            rango.min_col == 7
            and rango.max_col == 7
            and rango.max_row >= fila_inicio
            and rango.min_row < fila_total_original
        ):
            hoja.unmerge_cells(
                str(
                    rango
                )
            )

    fila_total = asegurar_filas_antes_de_total(
        hoja,
        fila_inicio,
        fila_total_original,
        len(
            fallas
        ),
        fila_inicio,
        range(
            2,
            8,
        ),
    )

    limpiar_datos_sin_tocar_combinadas(
        hoja,
        fila_inicio,
        fila_total - 1,
        2,
        7,
    )

    posiciones = defaultdict(
        list
    )

    for fila, falla in enumerate(
        fallas,
        start=fila_inicio,
    ):
        copiar_estilo_fila(
            hoja,
            fila_inicio,
            fila,
            range(
                2,
                8,
            ),
        )

        copiar_altura_fila(
            hoja,
            fila_inicio,
            fila,
        )

        hoja.cell(
            fila,
            2,
        ).value = falla[
            "falla_no"
        ]

        hoja.cell(
            fila,
            3,
        ).value = falla[
            "maquina"
        ]

        hoja.cell(
            fila,
            4,
        ).value = falla[
            "descripcion"
        ]

        hoja.cell(
            fila,
            5,
        ).value = falla[
            "minutos"
        ]

        hoja.cell(
            fila,
            6,
        ).value = falla[
            "eventos"
        ]

        posiciones[
            falla[
                "maquina"
            ]
        ].append(
            fila
        )

    for maquina, filas in posiciones.items():
        primera = filas[0]
        ultima = filas[-1]

        hoja.cell(
            primera,
            7,
        ).value = (
            f"=SUM(E{primera}:E{ultima})"
        )

        if ultima > primera:
            hoja.merge_cells(
                start_row=primera,
                start_column=7,
                end_row=ultima,
                end_column=7,
            )

    ultima_fila_datos = (
        fila_inicio
        + len(
            fallas
        )
        - 1
    )

    if not fallas:
        ultima_fila_datos = fila_inicio

    hoja.cell(
        fila_total,
        4,
    ).value = (
        "TOTAL TIEMPOS PERDIDOS AREA"
    )

    hoja.cell(
        fila_total,
        5,
    ).value = (
        f"=SUM(E{fila_inicio}:E{ultima_fila_datos})"
    )


def llenar_datos_pareto(
    hoja,
    fallas: list[dict[str, Any]],
) -> None:
    fila_inicio = 4
    fila_fin_base = 38
    fila_insercion = 39
    fila_total_base = 40

    ordenadas = sorted(
        fallas,
        key=lambda item:
            item[
                "minutos"
            ],
        reverse=True,
    )

    capacidad_base = (
        fila_fin_base
        - fila_inicio
        + 1
    )

    extra = max(
        0,
        len(
            ordenadas
        )
        - capacidad_base,
    )

    if extra > 0:
        hoja.insert_rows(
            fila_insercion,
            amount=extra,
        )

        for fila in range(
            fila_insercion,
            fila_insercion + extra,
        ):
            copiar_estilo_fila(
                hoja,
                fila_inicio,
                fila,
                range(
                    4,
                    9,
                ),
            )

            copiar_altura_fila(
                hoja,
                fila_inicio,
                fila,
            )

    fila_total = (
        fila_total_base
        + extra
    )

    fila_fin_datos = max(
        fila_fin_base + extra,
        fila_inicio,
    )

    limpiar_datos_sin_tocar_combinadas(
        hoja,
        fila_inicio,
        fila_fin_datos,
        4,
        8,
    )

    for numero, falla in enumerate(
        ordenadas,
        start=1,
    ):
        fila = (
            fila_inicio
            + numero
            - 1
        )

        copiar_estilo_fila(
            hoja,
            fila_inicio,
            fila,
            range(
                4,
                9,
            ),
        )

        copiar_altura_fila(
            hoja,
            fila_inicio,
            fila,
        )

        hoja.cell(
            fila,
            4,
        ).value = numero

        hoja.cell(
            fila,
            5,
        ).value = falla[
            "maquina"
        ]

        hoja.cell(
            fila,
            6,
        ).value = falla[
            "descripcion"
        ]

        hoja.cell(
            fila,
            7,
        ).value = falla[
            "minutos"
        ]

        hoja.cell(
            fila,
            8,
        ).value = falla[
            "eventos"
        ]

    ultima_fila_datos = (
        fila_inicio
        + len(
            ordenadas
        )
        - 1
    )

    if not ordenadas:
        ultima_fila_datos = fila_inicio

    hoja.cell(
        fila_total,
        6,
    ).value = (
        "TOTAL TIEMPOS PERDIDOS AREA"
    )

    hoja.cell(
        fila_total,
        7,
    ).value = (
        f"=SUM(G{fila_inicio}:G{ultima_fila_datos})"
    )

    hoja[
        "J4"
    ] = "PARETO 80/20"

    hoja[
        "J5"
    ] = "EVENTOS 20%"

    hoja[
        "K5"
    ] = (
        f"=COUNTA(D{fila_inicio}:D{ultima_fila_datos})*20%"
    )

    hoja[
        "J6"
    ] = "TIEMPO 80 %"

    hoja[
        "K6"
    ] = (
        f"=G{fila_total}*80%"
    )


def diligenciar_pareto(
    plantilla: Path,
    salida: Path,
    agregados: dict[
        str,
        dict[
            tuple[int, str],
            dict[str, Any],
        ],
    ],
    fallas_por_maquina: dict[
        str,
        list[dict[str, Any]],
    ],
) -> None:
    shutil.copy2(
        plantilla,
        salida,
    )

    excel, pythoncom = _excel_com()
    workbook = None

    try:
        workbook = excel.Workbooks.Open(
            str(
                salida.resolve()
            )
        )

        hoja_tiempos = _buscar_hoja_com(
            workbook,
            "TIEMPOS PERDIDOS",
        )

        _llenar_matriz_com(
            hoja_tiempos,
            agregados[
                "correctivo"
            ],
        )

        fallas = fallas_para_pareto(
            fallas_por_maquina
        )

        hoja_fallas = _buscar_hoja_com(
            workbook,
            "FALLAS TAPAS",
        )

        _llenar_fallas_tapas_com(
            hoja_fallas,
            fallas,
        )

        hoja_datos = _buscar_hoja_com(
            workbook,
            "DATOS PARETO",
        )

        _llenar_datos_pareto_com(
            hoja_datos,
            fallas,
        )

        workbook.Save()

    finally:
        if workbook is not None:
            workbook.Close(
                SaveChanges=True
            )

        _cerrar_excel_com(
            excel,
            pythoncom,
        )


def resumen_proceso(
    registros: list[dict[str, Any]],
) -> dict[str, Any]:
    resumen = {
        "total_registros":
            len(
                registros
            ),
        "correctivos":
            0,
        "preventivos":
            0,
        "cambios":
            0,
        "min_correctivo":
            0,
        "min_preventivo":
            0,
        "min_cambio":
            0,
    }

    for registro in registros:
        categoria = registro[
            "categoria"
        ]

        if categoria == "correctivo":
            resumen[
                "correctivos"
            ] += 1

            resumen[
                "min_correctivo"
            ] += registro[
                "minutos"
            ]

        elif categoria == "preventivo":
            resumen[
                "preventivos"
            ] += 1

            resumen[
                "min_preventivo"
            ] += registro[
                "minutos"
            ]

        elif categoria == "cambio":
            resumen[
                "cambios"
            ] += 1

            resumen[
                "min_cambio"
            ] += registro[
                "minutos"
            ]

    return resumen


def localizar_plantilla(
    nombres: list[str],
) -> Path:
    formatos = carpeta_formatos()

    for nombre in nombres:
        ruta = formatos / nombre

        if (
            ruta.exists()
            and not ruta.name.startswith(
                "~$"
            )
        ):
            return ruta

    archivos = [
        archivo
        for archivo in (
            list(
                formatos.glob(
                    "*.xlsx"
                )
            )
            + list(
                formatos.glob(
                    "*.xlsm"
                )
            )
        )
        if not archivo.name.startswith(
            "~$"
        )
    ]

    for archivo in archivos:
        limpio = normalizar(
            archivo.name
        )

        if any(
            normalizar(
                palabra
            ) in limpio
            for palabra in nombres
        ):
            return archivo

    raise FileNotFoundError(
        "No se encontró el formato requerido en la carpeta 'formatos'."
    )


@tiempos_perdidos_bp.route(
    "/",
    methods=[
        "GET",
        "POST",
    ],
)
def index():
    resultado = None

    if request.method == "POST":
        archivo = request.files.get(
            "archivo_entrega"
        )

        if (
            archivo is None
            or archivo.filename == ""
        ):
            flash(
                "Seleccione el archivo de Entrega de Turno.",
                "error",
            )

            return redirect(
                url_for(
                    "tiempos_perdidos.index"
                )
            )

        extension = Path(
            archivo.filename
        ).suffix.lower()

        if extension not in EXTENSIONES_PERMITIDAS:
            flash(
                "El archivo debe ser .xlsx o .xlsm.",
                "error",
            )

            return redirect(
                url_for(
                    "tiempos_perdidos.index"
                )
            )

        identificador = uuid.uuid4().hex[
            :10
        ]

        nombre_seguro = secure_filename(
            archivo.filename
        )

        ruta_entrada = (
            carpeta_uploads()
            / f"{identificador}_{nombre_seguro}"
        )

        archivo.save(
            ruta_entrada
        )

        try:
            lectura = leer_entrega_turno(
                ruta_entrada
            )

            registros = lectura[
                "registros"
            ]

            agregados = construir_agregados(
                registros
            )

            fallas_por_maquina = construir_fallas_correctivas(
                registros
            )


            plantilla_tiempos = localizar_plantilla(
                [
                    PLANTILLA_TIEMPOS,
                    "TIEMPOS JUNIO 2026 TAPA PLASTICA IBERPLAST.xlsx",
                    "TIEMPOS",
                ]
            )

            plantilla_pareto = localizar_plantilla(
                [
                    PLANTILLA_PARETO,
                    "PARETO 80-20 TAPAS JUNIO 2026.xlsx",
                    "PARETO",
                ]
            )

            mes = lectura.get(
                "mes"
            )

            anio = lectura.get(
                "anio"
            )

            sufijo = ""

            if mes and anio:
                sufijo = (
                    f"_{anio}_"
                    + calendar.month_name[
                        mes
                    ].upper()
                )

            salida_tiempos = (
                carpeta_salidas()
                / (
                    "TIEMPOS_TAPA_PLASTICA"
                    + sufijo
                    + "_DILIGENCIADO.xlsx"
                )
            )

            salida_pareto = (
                carpeta_salidas()
                / (
                    "PARETO_80_20_TAPAS"
                    + sufijo
                    + "_DILIGENCIADO.xlsx"
                )
            )

            diligenciar_formato_tiempos(
                plantilla_tiempos,
                salida_tiempos,
                agregados,
                fallas_por_maquina,
            )

            diligenciar_pareto(
                plantilla_pareto,
                salida_pareto,
                agregados,
                fallas_por_maquina,
            )

            resumen = resumen_proceso(
                registros
            )

            resultado = {
                **resumen,
                "hoja":
                    lectura[
                        "hoja"
                    ],
                "no_reconocidas":
                    lectura[
                        "maquinas_no_reconocidas"
                    ],
                "clases_ignoradas":
                    lectura[
                        "clases_ignoradas"
                    ],
                "archivo_tiempos":
                    salida_tiempos.name,
                "archivo_pareto":
                    salida_pareto.name,
            }

        except Exception as error:
            flash(
                str(error),
                "error",
            )

            return redirect(
                url_for(
                    "tiempos_perdidos.index"
                )
            )

        finally:
            try:
                ruta_entrada.unlink(
                    missing_ok=True
                )
            except OSError:
                pass

    return render_template(
        "tiempos_perdidos.html",
        resultado=resultado,
    )


@tiempos_perdidos_bp.route(
    "/descargar/<path:nombre>",
    methods=["GET"],
)
def descargar(nombre: str):
    nombre_archivo = Path(
        nombre
    ).name

    ruta = (
        carpeta_salidas()
        / nombre_archivo
    )

    if not ruta.exists():
        flash(
            "El archivo solicitado no existe.",
            "error",
        )

        return redirect(
            url_for(
                "tiempos_perdidos.index"
            )
        )

    return send_file(
        ruta,
        as_attachment=True,
        download_name=nombre_archivo,
    )
